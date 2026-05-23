"""
数据库模块 — SQLite存储 + 数据解析
从 fast_collector.py 提取已验证的 find_items() / normalize_item()
"""
import csv
import json
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import opencc
    _s2t = opencc.OpenCC('s2twp').convert   # 简体→台湾繁体(含词汇转换)
except ImportError:
    _s2t = None


def to_traditional(text):
    """简体→繁体, opencc不可用则原样返回"""
    if not text or not _s2t:
        return text or ''
    return _s2t(text)

DB_PATH = Path(__file__).parent / "xianyu_data.db"


def init_db(db_path=None):
    conn = sqlite3.connect(str(db_path or DB_PATH), check_same_thread=False)
    conn.execute('''CREATE TABLE IF NOT EXISTS products (
        item_id TEXT PRIMARY KEY,
        title TEXT, price TEXT, images TEXT, description TEXT,
        brief TEXT DEFAULT '',
        category TEXT, location TEXT, condition TEXT,
        seller_id TEXT, seller_nick TEXT, seller_avatar TEXT,
        want_count INTEGER DEFAULT 0, view_count INTEGER DEFAULT 0,
        post_time TEXT, url TEXT, collected_at TEXT, source TEXT
    )''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_seller ON products(seller_id)')
    # 迁移: 旧表可能没有 brief 列
    try:
        conn.execute('SELECT brief FROM products LIMIT 1')
    except sqlite3.OperationalError:
        conn.execute('ALTER TABLE products ADD COLUMN brief TEXT DEFAULT ""')
    # 迁移: 新增扩展分类+属性列
    _new_cols = [
        ('original_price', 'TEXT DEFAULT ""'),
        ('transport_fee', 'TEXT DEFAULT ""'),
        ('cat_dto_json', 'TEXT DEFAULT ""'),      # itemCatDTO完整JSON
        ('cpv_labels_json', 'TEXT DEFAULT ""'),    # cpvLabels完整JSON
        ('sold_count', 'INTEGER DEFAULT 0'),
        # 商品扩展
        ('item_status', 'TEXT DEFAULT ""'),        # 在线/已售出
        ('collect_count', 'INTEGER DEFAULT 0'),    # 收藏数
        ('favor_count', 'INTEGER DEFAULT 0'),      # 点赞数
        ('bargained', 'INTEGER DEFAULT 0'),        # 允许砍价
        ('item_tags', 'TEXT DEFAULT ""'),           # 标签JSON(包邮/保障等)
        ('promotion_tag', 'TEXT DEFAULT ""'),       # 促销标签
        ('gmt_create', 'TEXT DEFAULT ""'),          # 发布时间戳
        # 卖家扩展
        ('seller_unique_name', 'TEXT DEFAULT ""'),  # 卖家唯一名
        ('seller_signature', 'TEXT DEFAULT ""'),    # 卖家简介
        ('seller_sold_count', 'INTEGER DEFAULT 0'), # 历史成交数
        ('seller_item_count', 'INTEGER DEFAULT 0'), # 在售商品数
        ('seller_reg_days', 'INTEGER DEFAULT 0'),   # 注册天数
        ('seller_good_rate', 'TEXT DEFAULT ""'),     # 好评率
        ('seller_reply_rate', 'TEXT DEFAULT ""'),    # 24h回复率
        ('seller_reply_time', 'TEXT DEFAULT ""'),    # 平均回复时间
        ('seller_last_active', 'TEXT DEFAULT ""'),   # 最后活跃
        ('seller_playboy', 'INTEGER DEFAULT 0'),     # 玩家认证
        ('seller_zhima_auth', 'INTEGER DEFAULT 0'),  # 芝麻认证
        ('seller_zhima_level', 'TEXT DEFAULT ""'),   # 芝麻信用等级
        ('seller_good_remark', 'INTEGER DEFAULT 0'), # 好评数
        ('seller_bad_remark', 'INTEGER DEFAULT 0'),  # 差评数
        ('seller_city', 'TEXT DEFAULT ""'),           # 卖家城市
        # 圈子
        ('group_name', 'TEXT DEFAULT ""'),            # 所属圈子
        ('group_member_count', 'INTEGER DEFAULT 0'),  # 圈子人数
        # ── 第二批深挖字段 ──
        ('category_id', 'TEXT DEFAULT ""'),            # 顶层分类ID(categoryId)
        ('leaf_id', 'TEXT DEFAULT ""'),                # 叶子分类ID(leafId)
        ('item_label_texts', 'TEXT DEFAULT ""'),       # 分类标签文字
        ('sold_price', 'TEXT DEFAULT ""'),             # 成交价
        ('item_type', 'TEXT DEFAULT ""'),              # 商品类型
        ('gmt_create_str', 'TEXT DEFAULT ""'),         # 发布时间(可读)
        ('common_tags_text', 'TEXT DEFAULT ""'),       # 通用标签(包邮等)
        ('video_url', 'TEXT DEFAULT ""'),              # 视频链接
        ('trade_access_type', 'TEXT DEFAULT ""'),      # 交易方式
        ('seller_level', 'TEXT DEFAULT ""'),           # 卖家等级
        ('seller_portrait_url', 'TEXT DEFAULT ""'),    # 卖家头像URL
        ('seller_register_time', 'TEXT DEFAULT ""'),   # 卖家注册时间戳
        ('seller_yxp_pro', 'INTEGER DEFAULT 0'),       # 闲鱼Pro卖家
        ('seller_default_remark', 'INTEGER DEFAULT 0'),# 中评数
        ('seller_identity_tags', 'TEXT DEFAULT ""'),   # 身份认证标签
        ('seller_type', 'TEXT DEFAULT ""'),            # 卖家类型
        # 第三批: 分类属性详情
        ('label_props_detail', 'TEXT DEFAULT ""'),     # 标签属性详情(分类=xx|年代=xx|材质=xx)
    ]
    for col_name, col_def in _new_cols:
        try:
            conn.execute(f'SELECT {col_name} FROM products LIMIT 1')
        except sqlite3.OperationalError:
            conn.execute(f'ALTER TABLE products ADD COLUMN {col_name} {col_def}')
    conn.commit()
    return conn


def save_items(conn, items, source='http'):
    saved = 0
    for p in items:
        try:
            conn.execute('''INSERT OR IGNORE INTO products
                (item_id,title,price,images,description,brief,category,location,condition,
                 seller_id,seller_nick,seller_avatar,want_count,view_count,post_time,url,collected_at,source,
                 original_price,transport_fee,cat_dto_json,cpv_labels_json,sold_count)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (p.get('itemId', ''), p.get('title', ''), p.get('price', ''),
                 json.dumps(p.get('images', []), ensure_ascii=False),
                 p.get('description', ''), p.get('brief', ''),
                 p.get('category', ''), p.get('location', ''),
                 p.get('condition', ''), p.get('sellerId', ''), p.get('sellerNick', ''),
                 p.get('sellerAvatar', ''), p.get('wantCount', 0), p.get('viewCount', 0),
                 p.get('postTime', ''), p.get('url', ''),
                 datetime.now().isoformat(), source,
                 p.get('originalPrice', ''), p.get('transportFee', ''),
                 p.get('catDtoJson', ''), p.get('cpvLabelsJson', ''),
                 p.get('soldCount', 0)))
            if conn.execute('SELECT changes()').fetchone()[0] > 0:
                saved += 1
        except Exception:
            pass
    conn.commit()
    return saved


# ===== 递归item发现 (来自 fast_collector.py:113-140) =====
def find_items(node, depth=0, _parent_ctx=None):
    """递归查找item对象, 自动合并父级wrapper的scalar字段(保留状态/标签信息)"""
    if depth > 10 or not node:
        return []
    if isinstance(node, list):
        result = []
        for x in node:
            result.extend(find_items(x, depth + 1, _parent_ctx))
        return result
    if isinstance(node, dict):
        has_id = any(k in node for k in ['itemId', 'tradeItemId', 'id'])
        has_content = any(k in node for k in [
            'title', 'titleSummary', 'desc', 'picUrl', 'mainPic', 'mainPicInfo',
            'price', 'soldPrice', 'priceInfo', 'images',
        ])
        if has_id and has_content:
            has_rich = any(k in node for k in [
                'title', 'titleSummary', 'picUrl', 'images', 'mainPicInfo', 'mainPic',
            ])
            if has_rich:
                if _parent_ctx:
                    merged = dict(node)
                    for k, v in _parent_ctx.items():
                        if k not in merged:
                            merged[k] = v
                    return [merged]
                return [node]
        # 收集当前层的scalar字段作为子item的上下文
        ctx = dict(_parent_ctx) if _parent_ctx else {}
        for k, v in node.items():
            if isinstance(v, (str, int, float, bool)):
                ctx[k] = v
        result = []
        for v in node.values():
            if isinstance(v, (dict, list)):
                result.extend(find_items(v, depth + 1, ctx or None))
        return result
    return []


# ===== 字段标准化 (来自 fast_collector.py:143-235) =====

def normalize_price(price_str):
    """将价格转为纯数字: '1.10万' → '11000', '¥850' → '850'"""
    if not price_str:
        return ''
    s = str(price_str).strip()
    s = s.replace('¥', '').replace('￥', '').replace(',', '').replace(' ', '')
    multiplier = 1
    if '万' in s:
        s = s.replace('万', '')
        multiplier = 10000
    try:
        val = float(s) * multiplier
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}"
    except (ValueError, TypeError):
        return str(price_str)


def _extract_video_covers(raw):
    """从视频字段提取封面图URL (内聚清洗: strip/去空/补协议/保序去重)"""
    _COVER_KEYS = ('coverUrl', 'coverImage', 'firstFrameUrl', 'snapshotUrl',
                   'posterUrl', 'cover', 'imageUrl', 'url')
    covers = []
    # videoInfos (列表)
    vlist = raw.get('videoInfos', [])
    if isinstance(vlist, list):
        for vi in vlist:
            if isinstance(vi, dict):
                for key in _COVER_KEYS:
                    url = vi.get(key, '')
                    if url and isinstance(url, str):
                        covers.append(url)
                        break
    # videoInfo / videoDO / mainVideo / videoPlayInfo (单对象)
    for field in ('videoInfo', 'videoDO', 'mainVideo', 'videoPlayInfo'):
        vi = raw.get(field, {})
        if isinstance(vi, dict):
            for key in _COVER_KEYS:
                url = vi.get(key, '')
                if url and isinstance(url, str):
                    covers.append(url)
                    break
    # 内聚清洗: 去空/strip/补协议/保序去重
    seen = set()
    clean = []
    for url in covers:
        if not isinstance(url, str):
            continue
        url = url.strip()
        if not url:
            continue
        if url.startswith('//'):
            url = 'https:' + url
        if url not in seen:
            seen.add(url)
            clean.append(url)
    return clean


def normalize_item(raw):
    # exContent 合并
    exc = raw.get('exContent')
    if isinstance(exc, dict) and exc.get('itemId'):
        merged = dict(raw)
        for k, v in exc.items():
            if k not in merged or not merged[k]:
                merged[k] = v
        raw = merged

    user = raw.get('user', {}) if isinstance(raw.get('user'), dict) else {}
    dp = raw.get('detailParams', {}) if isinstance(raw.get('detailParams'), dict) else {}

    # images
    images = []
    if raw.get('images') and isinstance(raw['images'], list):
        images = [img if isinstance(img, str) else img.get('url', '') for img in raw['images']]
    elif raw.get('picUrl'):
        images = [raw['picUrl']]
    elif raw.get('mainPic'):
        images = [raw['mainPic']]
    elif raw.get('mainPicInfo') and isinstance(raw['mainPicInfo'], dict):
        images = [raw['mainPicInfo'].get('url', '')]
    if not images:
        pi = raw.get('picInfo')
        if isinstance(pi, dict) and pi.get('picUrl'):
            images = [pi['picUrl']]
    if not images and dp.get('picUrl'):
        images = [dp['picUrl']]
    # 视频封面 fallback
    if not images:
        images = _extract_video_covers(raw)

    # title
    title = raw.get('title', '')
    if not title:
        ts = raw.get('titleSummary')
        if isinstance(ts, dict):
            title = ts.get('text', '')
    if not title:
        title = dp.get('title', '')
    if not title:
        title = raw.get('desc', '')

    # price
    price = raw.get('soldPrice', '') or raw.get('price', '')
    if isinstance(price, list):
        price = ''.join(p.get('text', '') for p in price if isinstance(p, dict))
        price = price.replace('¥', '').strip()
    if not price:
        pi = raw.get('priceInfo')
        if isinstance(pi, dict):
            price = pi.get('price', pi.get('showPrice', ''))
    if not price:
        price = dp.get('soldPrice', '')
    price = normalize_price(price)

    # location
    location = (raw.get('city', '') or raw.get('area', '') or
                raw.get('location', '') or raw.get('divisionName', ''))

    # sellerId
    seller_id = ''
    for sid in [raw.get('userId'), raw.get('sellerId'), raw.get('ownerId'),
                user.get('userId'), dp.get('userId')]:
        if sid and str(sid).isdigit() and len(str(sid)) >= 5:
            seller_id = str(sid)
            break

    seller_nick = (raw.get('sellerNick') or raw.get('nick') or raw.get('userNickName')
                   or user.get('userNick') or dp.get('userNick') or '')
    seller_avatar = (raw.get('sellerAvatar') or raw.get('avatar') or raw.get('userAvatarUrl')
                     or user.get('avatar') or dp.get('userAvatarUrl') or '')

    return {
        'itemId': str(raw.get('itemId', raw.get('tradeItemId', raw.get('id', '')))),
        'title': title,
        'price': str(price),
        'images': images,
        'description': raw.get('desc', raw.get('description', '')),
        'category': raw.get('categoryName', raw.get('catName', raw.get('categoryId', ''))),
        'location': location,
        'condition': raw.get('stuffStatus', ''),
        'sellerId': seller_id,
        'sellerNick': seller_nick,
        'sellerAvatar': seller_avatar,
        'wantCount': int(raw.get('wantCount', 0) or 0),
        'viewCount': int(raw.get('viewCount', 0) or 0),
        'postTime': raw.get('publishTime', raw.get('gmtCreate', '')),
        'url': raw.get('detailUrl', raw.get('redirectUrl', raw.get('shareUrl', ''))),
    }


# ===== 导出 =====

# 导出列顺序 (与台湾平台模板一致)
EXPORT_COLS = [
    # 基础
    ('title', '標題'),
    ('brief', '商品簡述'),
    ('price', '起標價'),
    ('quantity', '數量'),
    ('description', '說明'),
    ('images', '圖片'),
    ('item_id', '商品條碼'),
    # 淘宝分类
    ('tb_cat_id', '淘宝分類ID'),
    ('tb_cat_name', '淘宝分類名稱'),
    # 闲鱼分类层级
    ('gf_cat_id', '闲鱼大类ID'),
    ('gf_cat_name_big', '闲鱼大类名'),
    ('gf_channel_cat_id', '闲鱼精准类ID'),
    ('gf_cat_name_detail', '闲鱼分类名'),
    ('gf_root_channel', '闲鱼根频道ID'),
    ('gf_level2_channel', '闲鱼二级频道ID'),
    ('gf_level3_channel', '闲鱼三级频道ID'),
    # 商品属性
    ('brand', '品牌'),
    ('stuff_status', '成色'),
    ('size_spec', '尺寸规格'),
    ('material', '材质'),
    ('cpv_all', '全部属性'),
    # 价格/交易
    ('original_price', '原价'),
    ('transport_fee', '运费'),
    ('sold_count', '已售数'),
    ('item_status', '商品状态'),
    ('collect_count', '收藏数'),
    ('favor_count', '点赞数'),
    ('bargained', '允许砍价'),
    ('item_tags', '商品标签'),
    ('promotion_tag', '促销标签'),
    ('gmt_create', '发布时间戳'),
    # 卖家信息
    ('seller_id', '卖家ID'),
    ('seller_nick', '卖家昵称'),
    ('seller_unique_name', '卖家唯一名'),
    ('seller_city', '卖家城市'),
    ('seller_signature', '卖家简介'),
    ('seller_sold_count', '卖家历史成交数'),
    ('seller_item_count', '卖家在售数'),
    ('seller_reg_days', '卖家注册天数'),
    ('seller_good_rate', '卖家好评率'),
    ('seller_reply_rate', '卖家24h回复率'),
    ('seller_reply_time', '卖家平均回复时间'),
    ('seller_last_active', '卖家最后活跃'),
    ('seller_playboy', '玩家认证'),
    ('seller_zhima_auth', '芝麻认证'),
    ('seller_zhima_level', '芝麻信用等级'),
    ('seller_good_remark', '好评数'),
    ('seller_bad_remark', '差评数'),
    # 圈子
    ('group_name', '所属圈子'),
    ('group_member_count', '圈子人数'),
    # 分类深挖
    ('category_id', '顶层分类ID'),
    ('leaf_id', '叶子分类ID'),
    ('item_label_texts', '分类标签文字'),
    ('label_props_detail', '标签属性详情'),
    ('tb_cat_path', '淘宝分类完整路径'),
    # 商品深挖
    ('sold_price', '成交价'),
    ('item_type', '商品类型'),
    ('gmt_create_str', '发布时间'),
    ('common_tags_text', '通用标签'),
    ('video_url', '视频链接'),
    ('trade_access_type', '交易方式'),
    # 卖家深挖
    ('seller_level', '卖家等级'),
    ('seller_portrait_url', '卖家头像'),
    ('seller_register_time', '卖家注册时间'),
    ('seller_yxp_pro', '闲鱼Pro卖家'),
    ('seller_default_remark', '中评数'),
    ('seller_identity_tags', '身份认证'),
    ('seller_type', '卖家类型'),
]

# ── 淘宝分类查找表 (tbCatId → {name, parent_cid}) ──
_TB_CATS = None

def _load_taobao_cats():
    """加载淘宝分类ID→名称+父级映射 (懒加载, 只读一次)"""
    global _TB_CATS
    if _TB_CATS is not None:
        return _TB_CATS
    _TB_CATS = {}
    tb_file = Path(__file__).parent / "taobao_itemcats.txt"
    if not tb_file.exists():
        tb_file = Path(r"C:/Users/USERNAME\taobao_itemcats.txt")
    if not tb_file.exists():
        return _TB_CATS
    try:
        with open(str(tb_file), "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.split(",", 3)
                if len(parts) >= 3:
                    cid = parts[0].strip()
                    parent_cid = parts[1].strip()
                    name = parts[2].strip().strip("'")
                    _TB_CATS[cid] = {"name": name, "parent_cid": parent_cid}
    except Exception:
        pass
    return _TB_CATS


def _get_taobao_path(tb_cats, cid):
    """获取淘宝分类完整路径 (根 > ... > 叶)"""
    path = []
    seen = set()
    current = cid
    while current and current != "0" and current not in seen:
        seen.add(current)
        info = tb_cats.get(current)
        if not info:
            break
        path.append(info["name"])
        current = info["parent_cid"]
    path.reverse()
    return " > ".join(path) if path else ""


def _build_image_paths(item_id, image_dir):
    """构建图片本地路径 (item_id文件夹), 管道符分隔"""
    if not image_dir or not item_id:
        return ''
    item_path = Path(image_dir) / str(item_id)
    if not item_path.exists():
        return ''

    files = sorted([
        fp for fp in item_path.iterdir()
        if fp.is_file() and fp.suffix.lower()
        in ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.heic', '.bmp')
    ])
    if not files:
        return ''
    return '|'.join(str(fp) for fp in files)


def _fmt_register_time(ts):
    """注册时间戳(毫秒) → 可读字符串"""
    if not ts:
        return ''
    try:
        t = int(ts)
        if t > 1000000000000:
            t = t // 1000
        from datetime import datetime as _dt
        return _dt.fromtimestamp(t).strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, OSError):
        return str(ts)


import re as _re
# Excel/openpyxl 不允许 \x00-\x08, \x0B-\x0C, \x0E-\x1F 等控制字符
_ILLEGAL_XML_RE = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _clean_for_xlsx(val):
    """清除Excel不允许的控制字符"""
    if isinstance(val, str):
        return _ILLEGAL_XML_RE.sub('', val)
    return val


def _format_row(item_id, title, price, images_json, description, brief, category, image_dir,
                 cat_dto_json='', cpv_labels_json='', original_price='', transport_fee='', sold_count=0,
                 item_status='', collect_count=0, favor_count=0, bargained=0,
                 item_tags='', promotion_tag='', gmt_create='',
                 seller_id='', seller_nick='',
                 seller_unique_name='', seller_city='', seller_signature='',
                 seller_sold_count=0, seller_item_count=0, seller_reg_days=0,
                 seller_good_rate='', seller_reply_rate='', seller_reply_time='',
                 seller_last_active='',
                 seller_playboy=0, seller_zhima_auth=0, seller_zhima_level='',
                 seller_good_remark=0, seller_bad_remark=0,
                 group_name='', group_member_count=0,
                 category_id='', leaf_id='', item_label_texts='',
                 sold_price='', item_type='', gmt_create_str='',
                 common_tags_text='', video_url='', trade_access_type='',
                 seller_level='', seller_portrait_url='', seller_register_time='',
                 seller_yxp_pro=0, seller_default_remark=0, seller_identity_tags='',
                 seller_type='', label_props_detail=''):
    """按 EXPORT_COLS 顺序格式化一行, 同时做简→繁转换"""
    # 图片
    if image_dir:
        img_text = _build_image_paths(item_id, image_dir)
    else:
        img_text = ''
        try:
            imgs = json.loads(images_json) if images_json else []
            if imgs:
                img_text = '|'.join(imgs)
        except (json.JSONDecodeError, TypeError):
            img_text = images_json or ''

    # 简→繁
    t_title = to_traditional(title or '')
    t_brief = to_traditional(brief or '')
    t_desc = to_traditional(description or '')

    # 价格转数字
    try:
        price_num = float(price) if price else 0
    except (ValueError, TypeError):
        price_num = 0

    # 拆分 category: "catId|channelCatId|名称|tbCatId"
    cat_parts = (category or '').split('|')
    gf_cat_id = cat_parts[0] if len(cat_parts) > 0 else ''
    gf_channel_cat_id = cat_parts[1] if len(cat_parts) > 1 else ''
    gf_cat_name_detail = cat_parts[2] if len(cat_parts) > 2 else ''
    tb_cat_id = cat_parts[3].strip() if len(cat_parts) > 3 else ''

    # 查淘宝分类
    tb_cats = _load_taobao_cats()
    tb_cat_path = _get_taobao_path(tb_cats, tb_cat_id) if tb_cat_id and tb_cats else ''
    tb_cat_id_num = int(tb_cat_id) if tb_cat_id and tb_cat_id.isdigit() else (tb_cat_id or '')

    # 闲鱼大类名 (catId也在淘宝分类库里)
    gf_cat_name_big = ''
    if gf_cat_id:
        info = tb_cats.get(gf_cat_id)
        if info:
            gf_cat_name_big = info['name']

    # 解析 itemCatDTO
    gf_root_channel = ''
    gf_level2_channel = ''
    gf_level3_channel = ''
    if cat_dto_json:
        try:
            cat_dto = json.loads(cat_dto_json)
            gf_root_channel = str(cat_dto.get('rootChannelCatId', ''))
            gf_level2_channel = str(cat_dto.get('level2ChannelCatId', ''))
            gf_level3_channel = str(cat_dto.get('level3ChannelCatId', ''))
        except (json.JSONDecodeError, TypeError):
            pass

    # 解析 cpvLabels
    brand = ''
    stuff_status = ''
    size_spec = ''
    material = ''
    cpv_all_parts = []
    if cpv_labels_json:
        try:
            cpv_list = json.loads(cpv_labels_json)
            for cpv in cpv_list:
                pname = cpv.get('propertyName', '')
                vname = cpv.get('valueName', '')
                if pname and vname:
                    cpv_all_parts.append(f"{pname}:{vname}")
                    if pname == '品牌':
                        brand = vname
                    elif pname == '成色':
                        stuff_status = vname
                    elif pname in ('尺寸', '尺码', '规格', '容量'):
                        size_spec = vname
                    elif pname == '材质':
                        material = vname
        except (json.JSONDecodeError, TypeError):
            pass
    cpv_all = ' | '.join(cpv_all_parts)

    # 原价/运费
    try:
        orig_price = float(original_price) if original_price else ''
    except (ValueError, TypeError):
        orig_price = original_price or ''
    try:
        t_fee = float(transport_fee) if transport_fee else ''
    except (ValueError, TypeError):
        t_fee = transport_fee or ''

    # 标签格式化
    tags_text = ''
    if item_tags:
        try:
            tags_list = json.loads(item_tags)
            tags_text = ' | '.join(tags_list) if isinstance(tags_list, list) else str(item_tags)
        except (json.JSONDecodeError, TypeError):
            tags_text = item_tags or ''

    # 发布时间格式化
    gmt_str = ''
    if gmt_create:
        try:
            ts = int(gmt_create)
            if ts > 1000000000000:
                ts = ts // 1000
            from datetime import datetime as _dt
            gmt_str = _dt.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, OSError):
            gmt_str = str(gmt_create)

    # 商品條碼轉純數字
    try:
        item_id_num = int(item_id) if item_id and str(item_id).isdigit() else (item_id or '')
    except (ValueError, TypeError):
        item_id_num = item_id or ''

    row = [
        t_title, t_brief, price_num, 1, t_desc, img_text, item_id_num,
        tb_cat_id_num, tb_cat_path,
        gf_cat_id, gf_cat_name_big,
        gf_channel_cat_id, gf_cat_name_detail,
        gf_root_channel, gf_level2_channel, gf_level3_channel,
        brand, stuff_status, size_spec, material, cpv_all,
        orig_price, t_fee, sold_count or 0,
        item_status or '', collect_count or 0, favor_count or 0,
        '是' if bargained else '否', tags_text, promotion_tag or '', gmt_str,
        seller_id or '', seller_nick or '', seller_unique_name or '',
        seller_city or '', seller_signature or '',
        seller_sold_count or 0, seller_item_count or 0, seller_reg_days or 0,
        seller_good_rate or '', seller_reply_rate or '', seller_reply_time or '',
        seller_last_active or '',
        '是' if seller_playboy else '否',
        '是' if seller_zhima_auth else '否',
        seller_zhima_level or '',
        seller_good_remark or 0, seller_bad_remark or 0,
        group_name or '', group_member_count or 0,
        category_id or '', leaf_id or '', item_label_texts or '',
        label_props_detail or '', tb_cat_path,
        sold_price or '', item_type or '', gmt_create_str or '',
        common_tags_text or '', video_url or '', trade_access_type or '',
        seller_level or '', seller_portrait_url or '',
        _fmt_register_time(seller_register_time),
        '是' if seller_yxp_pro else '否',
        seller_default_remark or 0,
        seller_identity_tags or '', seller_type or '',
    ]
    return [_clean_for_xlsx(v) for v in row]


_EXPORT_SQL = '''SELECT item_id, title, price, images, description,
    COALESCE(brief,""), COALESCE(category,""),
    COALESCE(cat_dto_json,""), COALESCE(cpv_labels_json,""),
    COALESCE(original_price,""), COALESCE(transport_fee,""), COALESCE(sold_count,0),
    COALESCE(item_status,""), COALESCE(collect_count,0), COALESCE(favor_count,0),
    COALESCE(bargained,0), COALESCE(item_tags,""), COALESCE(promotion_tag,""), COALESCE(gmt_create,""),
    seller_id, seller_nick,
    COALESCE(seller_unique_name,""), COALESCE(seller_city,""), COALESCE(seller_signature,""),
    COALESCE(seller_sold_count,0), COALESCE(seller_item_count,0), COALESCE(seller_reg_days,0),
    COALESCE(seller_good_rate,""), COALESCE(seller_reply_rate,""), COALESCE(seller_reply_time,""),
    COALESCE(seller_last_active,""),
    COALESCE(seller_playboy,0), COALESCE(seller_zhima_auth,0), COALESCE(seller_zhima_level,""),
    COALESCE(seller_good_remark,0), COALESCE(seller_bad_remark,0),
    COALESCE(group_name,""), COALESCE(group_member_count,0),
    COALESCE(category_id,""), COALESCE(leaf_id,""), COALESCE(item_label_texts,""),
    COALESCE(sold_price,""), COALESCE(item_type,""), COALESCE(gmt_create_str,""),
    COALESCE(common_tags_text,""), COALESCE(video_url,""), COALESCE(trade_access_type,""),
    COALESCE(seller_level,""), COALESCE(seller_portrait_url,""), COALESCE(seller_register_time,""),
    COALESCE(seller_yxp_pro,0), COALESCE(seller_default_remark,0), COALESCE(seller_identity_tags,""),
    COALESCE(seller_type,""), COALESCE(label_props_detail,"")
    FROM products ORDER BY collected_at DESC'''


def export_xlsx(conn, output_path=None, image_dir=None):
    """导出为 XLSX 格式 (精简9列: 標題~商品條碼 + 淘宝分類ID/名稱, 简→繁转换)"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    if output_path is None:
        output_path = Path(__file__).parent / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    rows = conn.execute(_EXPORT_SQL).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "商品数据"

    # 精简9列表头
    slim_headers = ['標題', '商品簡述', '起標價', '數量', '說明', '圖片',
                    '商品條碼', '淘宝分類ID', '淘宝分類名稱']
    ws.append(slim_headers)

    # 列宽
    slim_widths = [40, 30, 10, 6, 50, 80, 20, 14, 50]
    for i, w in enumerate(slim_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    for row in rows:
        (item_id, title, price, images_json, description, brief, category,
         cat_dto_json, cpv_labels_json, original_price, transport_fee, sold_count,
         item_status, collect_count, favor_count, bargained, item_tags, promotion_tag, gmt_create,
         seller_id, seller_nick,
         seller_unique_name, seller_city, seller_signature,
         seller_sold_count, seller_item_count, seller_reg_days,
         seller_good_rate, seller_reply_rate, seller_reply_time, seller_last_active,
         seller_playboy, seller_zhima_auth, seller_zhima_level,
         seller_good_remark, seller_bad_remark,
         group_name, group_member_count,
         category_id, leaf_id, item_label_texts,
         sold_price, item_type, gmt_create_str,
         common_tags_text, video_url, trade_access_type,
         seller_level, seller_portrait_url, seller_register_time,
         seller_yxp_pro, seller_default_remark, seller_identity_tags,
         seller_type, label_props_detail) = row
        full_row = _format_row(
            item_id, title, price, images_json, description, brief, category, image_dir,
            cat_dto_json, cpv_labels_json, original_price, transport_fee, sold_count,
            item_status, collect_count, favor_count, bargained, item_tags, promotion_tag, gmt_create,
            seller_id, seller_nick,
            seller_unique_name, seller_city, seller_signature,
            seller_sold_count, seller_item_count, seller_reg_days,
            seller_good_rate, seller_reply_rate, seller_reply_time, seller_last_active,
            seller_playboy, seller_zhima_auth, seller_zhima_level,
            seller_good_remark, seller_bad_remark,
            group_name, group_member_count,
            category_id, leaf_id, item_label_texts,
            sold_price, item_type, gmt_create_str,
            common_tags_text, video_url, trade_access_type,
            seller_level, seller_portrait_url, seller_register_time,
            seller_yxp_pro, seller_default_remark, seller_identity_tags,
            seller_type, label_props_detail)
        # 只取前9列: 標題(0), 商品簡述(1), 起標價(2), 數量(3), 說明(4), 圖片(5), 商品條碼(6), 淘宝分類ID(7), 淘宝分類名稱(8)
        ws.append(full_row[:9])

    # 商品條碼(G列)和淘宝分類ID(H列) 設為數字格式, 防止科學記號
    for col_letter in ['G', 'H']:
        for r in range(2, ws.max_row + 1):
            cell = ws[f'{col_letter}{r}']
            if cell.value is not None:
                cell.number_format = '0'

    wb.save(str(output_path))
    return len(rows), str(output_path)


def export_csv(conn, output_path=None, image_dir=None):
    """导出为 CSV 格式 (台湾平台列顺序, 简→繁转换)"""
    if output_path is None:
        output_path = Path(__file__).parent / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    rows = conn.execute(_EXPORT_SQL).fetchall()

    headers = [col[1] for col in EXPORT_COLS]

    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            (item_id, title, price, images_json, description, brief, category,
             cat_dto_json, cpv_labels_json, original_price, transport_fee, sold_count,
             item_status, collect_count, favor_count, bargained, item_tags, promotion_tag, gmt_create,
             seller_id, seller_nick,
             seller_unique_name, seller_city, seller_signature,
             seller_sold_count, seller_item_count, seller_reg_days,
             seller_good_rate, seller_reply_rate, seller_reply_time, seller_last_active,
             seller_playboy, seller_zhima_auth, seller_zhima_level,
             seller_good_remark, seller_bad_remark,
             group_name, group_member_count,
             category_id, leaf_id, item_label_texts,
             sold_price, item_type, gmt_create_str,
             common_tags_text, video_url, trade_access_type,
             seller_level, seller_portrait_url, seller_register_time,
             seller_yxp_pro, seller_default_remark, seller_identity_tags,
             seller_type, label_props_detail) = row
            w.writerow(_format_row(
                item_id, title, price, images_json, description, brief, category, image_dir,
                cat_dto_json, cpv_labels_json, original_price, transport_fee, sold_count,
                item_status, collect_count, favor_count, bargained, item_tags, promotion_tag, gmt_create,
                seller_id, seller_nick,
                seller_unique_name, seller_city, seller_signature,
                seller_sold_count, seller_item_count, seller_reg_days,
                seller_good_rate, seller_reply_rate, seller_reply_time, seller_last_active,
                seller_playboy, seller_zhima_auth, seller_zhima_level,
                seller_good_remark, seller_bad_remark,
                group_name, group_member_count,
                category_id, leaf_id, item_label_texts,
                sold_price, item_type, gmt_create_str,
                common_tags_text, video_url, trade_access_type,
                seller_level, seller_portrait_url, seller_register_time,
                seller_yxp_pro, seller_default_remark, seller_identity_tags,
                seller_type, label_props_detail))

    return len(rows), str(output_path)


def get_stats(conn, image_dir=None) -> dict:
    total = conn.execute('SELECT COUNT(*) FROM products').fetchone()[0]
    sellers = conn.execute('SELECT COUNT(DISTINCT seller_id) FROM products').fetchone()[0]
    sources = conn.execute('SELECT source, COUNT(*) FROM products GROUP BY source').fetchall()
    with_images = conn.execute("SELECT COUNT(*) FROM products WHERE images != '[]' AND images != ''").fetchone()[0]
    with_title = conn.execute("SELECT COUNT(*) FROM products WHERE title != ''").fetchone()[0]

    # 统计已下载图片的商品数
    downloaded_images = 0
    if image_dir:
        img_path = Path(image_dir)
        if img_path.exists():
            rows = conn.execute("SELECT item_id FROM products").fetchall()
            for (item_id,) in rows:
                item_dir = img_path / str(item_id)
                if item_dir.exists() and any(
                    fp.is_file() and fp.suffix.lower()
                    in ('.jpg', '.jpeg', '.png', '.webp', '.gif', '.heic', '.bmp')
                    for fp in item_dir.iterdir()
                ):
                    downloaded_images += 1

    return {
        'total': total,
        'sellers': sellers,
        'sources': dict(sources),
        'with_images': with_images,
        'with_title': with_title,
        'downloaded_images': downloaded_images,
    }
