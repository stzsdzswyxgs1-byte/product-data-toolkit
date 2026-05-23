# -*- coding: utf-8 -*-
"""
分類映射處理器
- 煤爐: 讀取映射表 (煤爐ID→奇摩ID), 標紅行拆到不合格
- 鹹魚: 讀取規則表 (關鍵詞→奇摩ID), 逐條匹配
"""
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Tuple, Optional, Callable
from processors.utils import append_reason

LogFn = Callable[[str], None]

# ─── 紅色檢測閾值 ───
RED_THRESHOLD = 0xCC  # R通道 > 0xCC 且 G/B < 0x66 視為紅色


def _parse_color(color_str: Optional[str]) -> Tuple[int, int, int]:
    """解析 openpyxl 顏色字串 (如 'FFFF0000') → (R, G, B)"""
    if not color_str or color_str == '00000000':
        return (0, 0, 0)
    s = color_str.lstrip('#')
    if len(s) == 8:  # AARRGGBB
        s = s[2:]    # 去掉 Alpha
    if len(s) != 6:
        return (0, 0, 0)
    try:
        return (int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16))
    except ValueError:
        return (0, 0, 0)


def _is_red_fill(cell) -> bool:
    """判斷單元格是否有紅色背景填充"""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    fg = fill.fgColor
    if fg and fg.rgb and fg.rgb != '00000000':
        r, g, b = _parse_color(str(fg.rgb))
        if r > RED_THRESHOLD and g < 0x66 and b < 0x66:
            return True
    return False


def load_mercari_mapping(mapping_path: str, log_fn: LogFn = print
                         ) -> Tuple[Dict[int, Tuple[str, str]], set]:
    """
    讀取煤爐映射奇摩分類.xlsx, 返回:
      map_dict: {煤爐ID: (奇摩ID_str, 奇摩分類_str)}
      red_ids:  set of 標紅的煤爐ID
    """
    log_fn(f"讀取映射表: {mapping_path}")
    wb = load_workbook(mapping_path, read_only=False, data_only=True)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    col_idx = {}
    for need in ['煤爐ID', '奇摩ID', '奇摩分類']:
        if need in headers:
            col_idx[need] = headers.index(need) + 1
        else:
            raise ValueError(f"映射表缺少欄位: {need}")

    map_dict = {}
    red_ids = set()
    total = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        try:
            mid_cell = ws.cell(row=row, column=col_idx['煤爐ID'])
            mid = int(float(str(mid_cell.value)))
            yid = str(int(float(str(ws.cell(row=row, column=col_idx['奇摩ID']).value))))
            yname = str(ws.cell(row=row, column=col_idx['奇摩分類']).value or '')

            map_dict[mid] = (yid, yname)

            # 檢查該行第一個有內容的cell是否標紅
            if _is_red_fill(mid_cell):
                red_ids.add(mid)
        except (ValueError, TypeError):
            continue

    wb.close()
    log_fn(f"映射表載入完成: {len(map_dict)} 條映射, {len(red_ids)} 條標紅")
    return map_dict, red_ids


def apply_mercari_mapping(df: pd.DataFrame, map_dict: dict, red_ids: set,
                          log_fn: LogFn = print) -> dict:
    """
    對 DataFrame 應用煤爐映射, 填入 拍賣類別/拍賣類別名稱,
    標紅的寫入 _filter_reason
    返回統計 dict
    """
    stats = {'mapped': 0, 'red': 0, 'unmapped': 0}

    cat_id_col = '_source_category_id'
    if cat_id_col not in df.columns:
        log_fn("[警告] 無 _source_category_id 欄位, 跳過映射")
        return stats

    for idx in df.index:
        raw = df.at[idx, cat_id_col]
        try:
            sid = int(float(str(raw)))
        except (ValueError, TypeError):
            df.at[idx, '_filter_reason'] = append_reason(
                df.at[idx, '_filter_reason'], '分類ID為空')
            stats['unmapped'] += 1
            continue

        if sid in map_dict:
            yid, yname = map_dict[sid]
            df.at[idx, '拍賣類別'] = yid
            df.at[idx, '拍賣類別名稱'] = yname

            if sid in red_ids:
                df.at[idx, '_filter_reason'] = append_reason(
                    df.at[idx, '_filter_reason'], '標紅分類')
                stats['red'] += 1
            else:
                stats['mapped'] += 1
        else:
            df.at[idx, '_filter_reason'] = append_reason(
                df.at[idx, '_filter_reason'], f'未映射分類({sid})')
            stats['unmapped'] += 1

    log_fn(f"映射完成: 合格{stats['mapped']} | 標紅{stats['red']} | 未映射{stats['unmapped']}")
    return stats


# ─── 鹹魚映射表分類 (淘宝ID→奇摩ID, 類似煤爐) ───

def load_goofish_mapping(mapping_path: str, log_fn: LogFn = print
                         ) -> Tuple[Dict[str, Tuple[str, str]], set]:
    """
    讀取鹹魚映射奇摩分類.xlsx (與煤爐映射邏輯一致), 返回:
      map_dict: {淘宝分類ID_str: (奇摩ID_str, 奇摩分類_str)}
      red_ids:  set of 標紅的淘宝分類ID_str
    欄位: 淘宝分类ID, 淘宝完整路径, 奇摩ID, 奇摩完整分類
    """
    log_fn(f"讀取鹹魚映射表: {mapping_path}")
    wb = load_workbook(mapping_path, read_only=False, data_only=True)
    ws = wb.active

    # 自動偵測欄位位置
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    col_src = col_kid = col_kname = None
    for i, h in enumerate(headers):
        if h in ('淘宝分类ID', '淘宝分類ID'):
            col_src = i + 1
        elif h == '奇摩ID':
            col_kid = i + 1
        elif h in ('奇摩完整分類', '奇摩分類'):
            col_kname = i + 1

    if not col_src or not col_kid:
        wb.close()
        raise ValueError(f"映射表缺少必要欄位 (需要: 淘宝分类ID, 奇摩ID). 現有: {headers}")

    map_dict = {}
    red_ids = set()

    for row in range(2, ws.max_row + 1):
        try:
            src_cell = ws.cell(row=row, column=col_src)
            src_val = src_cell.value
            if src_val is None:
                continue

            src_id = str(int(float(str(src_val))))
            kid_val = ws.cell(row=row, column=col_kid).value
            kname_val = ws.cell(row=row, column=col_kname).value if col_kname else ''

            kid = str(int(float(str(kid_val)))) if kid_val else ''
            kname = str(kname_val or '')

            if src_id and kid:
                map_dict[src_id] = (kid, kname)

            # 檢查淘宝分類ID欄位是否標紅 (與煤爐一致)
            if _is_red_fill(src_cell):
                red_ids.add(src_id)
        except (ValueError, TypeError):
            continue

    wb.close()
    log_fn(f"鹹魚映射表載入完成: {len(map_dict)} 條映射, {len(red_ids)} 條標紅")
    return map_dict, red_ids


def apply_goofish_mapping(df: pd.DataFrame, map_dict: dict, red_ids: set,
                          log_fn: LogFn = print) -> dict:
    """
    對 DataFrame 應用鹹魚映射 (淘宝分類ID → 奇摩ID),
    填入 拍賣類別/拍賣類別名稱, 標紅的寫入 _filter_reason
    返回統計 dict (與煤爐 apply_mercari_mapping 邏輯一致)
    """
    stats = {'mapped': 0, 'red': 0, 'unmapped': 0}

    cat_id_col = '_source_category_id'
    if cat_id_col not in df.columns:
        log_fn("[警告] 無 _source_category_id 欄位, 跳過映射")
        return stats

    for idx in df.index:
        raw = df.at[idx, cat_id_col]
        sid = str(raw).strip() if pd.notna(raw) else ''

        # 去掉小數點
        if '.' in sid:
            sid = sid.split('.')[0]

        if not sid or sid == 'nan':
            df.at[idx, '_filter_reason'] = append_reason(
                df.at[idx, '_filter_reason'], '淘宝分類ID為空')
            stats['unmapped'] += 1
            continue

        if sid in map_dict:
            yid, yname = map_dict[sid]
            df.at[idx, '拍賣類別'] = yid
            df.at[idx, '拍賣類別名稱'] = yname

            if sid in red_ids:
                df.at[idx, '_filter_reason'] = append_reason(
                    df.at[idx, '_filter_reason'], '標紅分類')
                stats['red'] += 1
            else:
                stats['mapped'] += 1
        else:
            df.at[idx, '_filter_reason'] = append_reason(
                df.at[idx, '_filter_reason'], f'未映射分類({sid})')
            stats['unmapped'] += 1

    log_fn(f"鹹魚映射完成: 合格{stats['mapped']} | 標紅{stats['red']} | 未映射{stats['unmapped']}")
    return stats


# ─── 規則分類 (鹹魚用, 舊版關鍵詞匹配) ───

def load_rules(rules_path: str, log_fn: LogFn = print) -> list:
    """
    讀取 規則.xlsx, 返回規則列表:
    [{'cat_id': str, 'cat_name': str, 'keywords': [str], 'excludes': [str]}, ...]
    """
    log_fn(f"讀取規則表: {rules_path}")
    df = pd.read_excel(rules_path, engine='openpyxl')

    needed = ['分類編號', '分類名稱', '規則']
    miss = [c for c in needed if c not in df.columns]
    if miss:
        raise ValueError(f"規則表缺少欄位: {miss}")

    rules = []
    for _, row in df.iterrows():
        cat_id = row.get('分類編號')
        cat_name = row.get('分類名稱', '')
        rule_str = row.get('規則', '')
        excl_str = row.get('不含關鍵詞', '')

        if pd.isna(cat_id) or pd.isna(rule_str) or not str(rule_str).strip():
            continue

        keywords = [k.strip() for k in str(rule_str).split(',') if k.strip()]
        excludes = []
        if pd.notna(excl_str) and str(excl_str).strip():
            excludes = [e.strip() for e in str(excl_str).split(',') if e.strip()]

        rules.append({
            'cat_id': str(int(float(str(cat_id)))),
            'cat_name': str(cat_name) if pd.notna(cat_name) else '',
            'keywords': keywords,
            'excludes': excludes,
        })

    log_fn(f"規則表載入完成: {len(rules)} 條規則")
    return rules


def apply_rules_classification(df: pd.DataFrame, rules: list,
                               log_fn: LogFn = print) -> dict:
    """
    對 DataFrame 逐行用規則匹配, 填入 拍賣類別/拍賣類別名稱
    返回統計 dict
    """
    stats = {'classified': 0, 'unclassified': 0}

    for idx in df.index:
        title = str(df.at[idx, '標題']) if pd.notna(df.at[idx, '標題']) else ''
        if not title:
            stats['unclassified'] += 1
            continue

        matched = False
        for rule in rules:
            # AND: 所有關鍵詞都要在標題中
            if all(kw in title for kw in rule['keywords']):
                # 排除詞: 任一出現就跳過
                if rule['excludes'] and any(ex in title for ex in rule['excludes']):
                    continue
                df.at[idx, '拍賣類別'] = rule['cat_id']
                df.at[idx, '拍賣類別名稱'] = rule['cat_name']
                matched = True
                stats['classified'] += 1
                break

        if not matched:
            df.at[idx, '_filter_reason'] = append_reason(
                df.at[idx, '_filter_reason'], '未分類')
            stats['unclassified'] += 1

    log_fn(f"規則分類完成: 已分類{stats['classified']} | 未分類{stats['unclassified']}")
    return stats


