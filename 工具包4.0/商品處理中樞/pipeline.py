# -*- coding: utf-8 -*-
"""
商品處理管線 — 核心調度器
負責串聯所有處理步驟, 根據配置和來源類型調度
"""
import json
import os
import re
import sys
import time
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Callable, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from adapters.mercari_adapter import adapt_mercari
from adapters.goofish_adapter import adapt_goofish
from processors.category_mapper import (
    load_mercari_mapping, apply_mercari_mapping,
    load_goofish_mapping, apply_goofish_mapping,
    load_rules, apply_rules_classification,
)
from processors.text_replacer import load_replacements, apply_replacements
from processors.price_cleaner import load_price_cleaner, apply_price_cleaning
from processors.keyword_filter import load_keywords, apply_keyword_filter
from processors.price_converter import apply_price_conversion
from processors.defaults_filler import apply_defaults
from processors.title_dedup import apply_dedup
from processors.desc_appender import apply_desc_append
from processors.auto_mapper import auto_map_missing
from processors.image_optimizer import optimize_first_images
from processors.seo_v64_full import run_v64_full
from processors.image_dewatermark import try_dewatermark_rejected
from checkpoint_manager import (
    CheckpointManager, APIMonitor, PauseException, set_monitor, get_monitor,
)
from quota_client import (
    QuotaClient, QuotaError, QuotaConfigError, QuotaExceeded, QuotaNetworkError,
)

LogFn = Callable[[str], None]

# 最終輸出列順序 (與自動刊登 test 格式一致)
OUTPUT_COLUMNS = [
    '標題', '商品簡述', '標籤', '起標價', '數量', '說明', '圖片',
    '商品條碼', '所在地', '商品類型', '商品狀況', '交貨方式',
    '付款方式', '出貨日期', '上架類型', '拍賣類別', '拍賣類別名稱',
]


class Pipeline:
    """商品處理管線"""

    def __init__(self, config: dict, log_fn: LogFn = print):
        self.config = config
        self._gui_log = log_fn  # GUI 顯示用 (精簡)
        self._detail_fp = None  # 詳細日誌檔案 (詳細)
        self.stats = {}
        self._stop = False
        # ★ Checkpoint: pipeline 開跑時建立, paused/crash 後 resume 用
        self.ckpt: 'CheckpointManager | None' = None
        self.monitor: 'APIMonitor | None' = None
        self._resume_mode = False  # True 時跳過 already-done stages
        self._paused_during_run = False

    def log(self, msg):
        """精簡 log: 同時寫 GUI 和詳細日誌"""
        self._gui_log(msg)
        if self._detail_fp:
            try:
                self._detail_fp.write(msg + '\n')
                self._detail_fp.flush()
            except: pass

    def log_detail(self, msg):
        """詳細 log: 只寫到日誌檔, GUI 不顯示 (避免刷屏)"""
        if self._detail_fp:
            try:
                self._detail_fp.write(msg + '\n')
                self._detail_fp.flush()
            except: pass

    def log_admin(self, title: str, fields: dict):
        """admin 維護專用結構化區塊 (給 admin Claude 診斷用).
        寫到詳細日誌, GUI 不顯示. 用 ASCII frame 容易 parse + grep.
        ★ 4.0.24+: 同事有問題把詳細日誌整檔發給 admin, admin 找 [ADMIN_*] 區塊看狀態."""
        if not self._detail_fp:
            return
        try:
            sep = '=' * 60
            self._detail_fp.write(f'\n{sep}\n[ADMIN_{title}]\n{sep}\n')
            for k, v in fields.items():
                # 多行 value 縮排
                if isinstance(v, str) and '\n' in v:
                    self._detail_fp.write(f'{k}:\n')
                    for line in v.split('\n'):
                        self._detail_fp.write(f'    {line}\n')
                else:
                    self._detail_fp.write(f'{k}: {v}\n')
            self._detail_fp.write(f'{sep}\n')
            self._detail_fp.flush()
        except Exception:
            pass

    def _dump_admin_summary(self, exit_status: str, **extra):
        """跑批結束 (success / paused / crashed) 都呼叫. 給 admin Claude 診斷用結尾摘要."""
        import datetime
        try:
            elapsed = (__import__('time').time() - self._fp_t_start) if self._fp_t_start else 0
        except Exception:
            elapsed = 0
        fields = {
            'ts': datetime.datetime.now().isoformat(timespec='seconds'),
            'exit_status': exit_status,  # completed / paused / crashed / user_stopped
            'total_runtime_s': f'{elapsed:.0f}',
            'batch_id': self._fp_batch_id or '?',
            'tg_id': self.config.get('tg_id', '(空)'),
            'input_rows': self._fp_input_rows or 0,
            'quota_consumed': self._quota_consumed or 0,
        }
        # APIMonitor 統計 — 看 VPN 抖動 / 真 API failure 比例
        # ★ 4.0.27: getattr 兼容舊 monitor (沒 total_* 欄位的客戶端)
        try:
            mon = self.monitor
            if mon:
                fields['api_monitor'] = (
                    f'consecutive_fails={getattr(mon, "consecutive_fails", "?")}, '
                    f'paused={getattr(mon, "_paused", "?")}, '
                    f'total_successes={getattr(mon, "total_successes", "(舊 monitor 無此欄)")}, '
                    f'total_real_failures={getattr(mon, "total_real_failures", "(舊)")}, '
                    f'total_network_blips={getattr(mon, "total_network_blips", "(舊)")}, '
                    f'total_pause_codes={getattr(mon, "total_pause_codes", "(舊)")}, '
                    f'total_skip_codes={getattr(mon, "total_skip_codes", "(舊)")}'
                )
        except Exception:
            pass
        # stats (各 stage 結果)
        if self.stats:
            stats_str_parts = []
            for k, v in self.stats.items():
                if isinstance(v, dict):
                    stats_str_parts.append(f'  {k}: {v}')
                else:
                    stats_str_parts.append(f'  {k}: {v}')
            fields['stats_per_stage'] = '\n'.join(stats_str_parts)
        # extra fields (split result 等)
        for k, v in extra.items():
            fields[k] = v
        self.log_admin('SUMMARY', fields)

    def _collect_admin_fingerprint(self) -> dict:
        """收集啟動環境 fingerprint — 給 admin Claude 看版本/GPU/網路狀態."""
        import sys, platform, datetime
        fp = {}
        fp['ts'] = datetime.datetime.now().isoformat(timespec='seconds')
        # 版本
        try:
            with open(os.path.join(os.path.dirname(__file__), 'current_version.txt'), encoding='utf-8') as f:
                fp['hub_version'] = f.read().strip()
        except Exception:
            fp['hub_version'] = '?'
        fp['python'] = sys.version.split()[0]
        fp['os'] = platform.platform()
        # torch / GPU
        try:
            import torch
            fp['torch'] = torch.__version__
            if torch.cuda.is_available():
                cap = torch.cuda.get_device_capability(0)
                props = torch.cuda.get_device_properties(0)
                fp['gpu'] = f'{props.name} (sm_{cap[0]}.{cap[1]}, {props.total_memory/(1024**3):.1f}GB)'
                fp['arch_list'] = str(torch.cuda.get_arch_list())
            else:
                fp['gpu'] = '(無 CUDA)'
        except Exception as e:
            fp['torch'] = f'IMPORT FAIL: {type(e).__name__}: {str(e)[:80]}'
        # numpy / pandas / Pillow / opencv / lama
        _name_map = {'cv2': 'opencv', 'PIL': 'pillow'}
        for mod_name in ('numpy', 'pandas', 'cv2', 'PIL', 'simple_lama_inpainting'):
            display = _name_map.get(mod_name, mod_name)
            try:
                mod = __import__(mod_name)
                fp[display] = getattr(mod, '__version__', '(no __version__)')
            except Exception:
                fp[display] = 'NOT INSTALLED'
        # config + 跑批 context
        cfg = self.config or {}
        fp['tg_id'] = cfg.get('tg_id', '(空)')
        fp['source_type'] = self._fp_source_type or '?'
        fp['process_mode'] = cfg.get('process_mode', 'raw')
        fp['output_dir'] = cfg.get('output_dir', '?')
        fp['input_file'] = self._fp_input_file or '?'
        fp['input_rows'] = self._fp_input_rows or 0
        fp['batch_id'] = self._fp_batch_id or '?'
        steps_cfg = cfg.get('steps', {})
        fp['enabled_steps'] = ','.join(k for k, v in steps_cfg.items() if v.get('enabled'))
        # 中介 ping (best-effort, 1 秒 timeout)
        try:
            import urllib.request, time as _t
            qcfg = cfg.get('quota', {})
            ep = qcfg.get('endpoint', '')
            if ep:
                t0 = _t.time()
                req = urllib.request.Request(ep.rstrip('/') + '/v1/health',
                                             headers={'User-Agent': 'hub-fp/4.0'})
                with urllib.request.urlopen(req, timeout=2) as r:
                    r.read(100)
                fp['mw_ping_ms'] = int((_t.time() - t0) * 1000)
            else:
                fp['mw_ping_ms'] = '(no endpoint)'
        except Exception as e:
            fp['mw_ping_ms'] = f'FAIL: {type(e).__name__}: {str(e)[:50]}'
        # 配額
        try:
            qc = self._quota_client
            if qc:
                fp['quota_used_at_start'] = self._fp_quota_at_start or '(未查)'
        except Exception:
            pass
        return fp

    def _open_detail_log(self, output_dir):
        """開啟詳細日誌檔"""
        import time as _t
        ts = _t.strftime('%Y%m%d_%H%M%S')
        path = os.path.join(output_dir, f'詳細日誌_{ts}.txt')
        try:
            os.makedirs(output_dir, exist_ok=True)
            self._detail_fp = open(path, 'w', encoding='utf-8')
            self._detail_path = path  # ★ 4.0.43: 存路徑給 run() 結尾 push 用
            self._detail_finalized = False  # ★ 4.0.44: idempotent flag
            # ★ 4.0.44 Bug 1: crash path 兜底 — 用戶關 GUI / process 結束時補 push
            #   normal flow (complete / pause) 已先跑 _finalize, idempotent flag 保證不重複.
            #   crash 時主流程 raise → 沒走 _finalize → atexit 兜這次.
            try:
                import atexit as _atexit
                # sync=True: process 退出時同步 push, daemon thread 來不及保住
                _atexit.register(self._finalize_detail_log, sync=True)
            except Exception: pass
            self._gui_log(f'[詳細日誌] 寫到: {path}')
            return path
        except Exception as e:
            self._gui_log(f'[詳細日誌] 開檔失敗: {e}')
            self._detail_path = None
            self._detail_finalized = True  # 開檔失敗就視同 finalized (不要再嘗試 push)
            return None

    def _close_detail_log(self):
        if self._detail_fp:
            try: self._detail_fp.close()
            except: pass
            self._detail_fp = None

    def _finalize_detail_log(self, sync=False):
        """★ 4.0.44: close + push 詳細日誌. Idempotent (只跑一次, run() / pause / crash / atexit 都呼叫).
        sync=False: background thread, 不阻塞 batch 結尾 GUI (normal flow 用)
        sync=True:  同步 push, 阻塞到完成 (atexit 兜底時用 — daemon thread 在 process 退出時來不及)
        ★ 4.0.61 fix: TG_USER_REDACTED 用戶 lama events 上來但 detail log 沒上來 — atexit 兜底失效.
           原因: normal flow set _detail_finalized=True 後啟 background thread, 用戶馬上關 GUI →
           process 退出 → atexit 跑 sync 路徑看 idempotent flag 直接 return →
           daemon thread 被 kill, push 沒完成. Fix: atexit sync 路徑不只看 flag,
           還要等 background thread 跑完才放行.
        """
        already_finalized = getattr(self, '_detail_finalized', True)
        if not already_finalized:
            self._detail_finalized = True
            # 1. close 檔 (確保 buffer flush 到檔)
            self._close_detail_log()
            # 2. push
            detail_path = getattr(self, '_detail_path', None)
            if not detail_path or not os.path.isfile(detail_path):
                return
            if not (hasattr(self, '_feedback_collector') and self._feedback_collector):
                return
            coll = self._feedback_collector

            def _safe_log(m, level='info'):
                try: self.log(m)
                except Exception: pass

            def _push_worker():
                try:
                    coll.push_detail_log(detail_path, log_fn=_safe_log)
                except Exception:
                    pass  # fire-and-forget, 永不擾用戶

            if sync:
                # atexit 兜底: 同步等 push 完, process 才退出
                _push_worker()
                return
            # normal flow: background thread daemon
            # ★ 4.0.61: 記住 thread 物件給 atexit 兜底用 (process 退出前等它跑完)
            try:
                import threading as _th
                t = _th.Thread(target=_push_worker, daemon=True, name='detail-log-push')
                self._detail_push_thread = t
                t.start()
            except Exception:
                pass  # 連 thread 都建不了就放棄
            return

        # ★ 4.0.61: 已 finalize 路徑 — atexit sync 兜底時要等 background thread 跑完
        if sync:
            t = getattr(self, '_detail_push_thread', None)
            if t and t.is_alive():
                try:
                    t.join(timeout=15)  # 給 push 最多 15s 跑完, process 才退
                except Exception:
                    pass

    def stop(self):
        self._stop = True
        # ★ 同時通知 monitor → 各個長階段的 batch loop 看 user_stop 提早結束
        try:
            from checkpoint_manager import get_monitor
            get_monitor().user_stop = True
        except Exception:
            pass

    def _check_stop(self):
        if self._stop:
            raise InterruptedError("用戶中止處理")

    def run(self, input_path: str, source_type: str,
            progress_fn: Optional[Callable] = None,
            resume: bool = False) -> dict:
        """
        執行完整管線
        input_path: 輸入 Excel 路徑
        source_type: 'mercari' 或 'goofish'
        progress_fn: callback(step_name, current, total)
        resume: True = 從 _checkpoint.* 接續跑
        返回: {'good_path': str, 'bad_path': str, 'uncat_path': str, 'stats': dict, 'paused': bool}
        """
        self._stop = False
        self._paused_during_run = False
        self._expensive_started = False  # 貴功能 (Step F-V65) 開跑後 → True; 用於決定 abort 退費
        self._quota_client = None
        self._quota_consumed = 0
        self.stats = {}
        # ★ 4.0.24: admin fingerprint 累積欄位 — 啟動 + 結尾 dump 給 admin Claude 看
        self._fp_source_type = source_type
        self._fp_input_file = os.path.basename(input_path) if input_path else ''
        self._fp_input_rows = 0
        self._fp_batch_id = ''
        self._fp_quota_at_start = ''
        self._fp_t_start = time.time() if 'time' in dir() else __import__('time').time()
        self._fp_stage_timing = {}  # stage_name → seconds
        self._fp_network_blip_count = 0  # network_blip 累積 (4.0.24 不算 monitor 但要記給 admin 看)
        total_steps = sum(1 for s in self.config.get('steps', {}).values() if s.get('enabled'))
        current_step = 0

        def _progress(name):
            nonlocal current_step
            current_step += 1
            if progress_fn:
                progress_fn(name, current_step, total_steps)

        # ★ 開啟詳細日誌 (寫到 output_dir, GUI 不顯示細節避免刷屏)
        output_dir = self.config.get('output_dir', '.')
        self._open_detail_log(output_dir)

        # ★ 4.0.35: reset adaptive cap registry — 每 batch 開始 fresh state
        try:
            from processors.utils import reset_adaptive_caps
            reset_adaptive_caps()
        except Exception:
            pass

        # ★ Checkpoint + APIMonitor 初始化
        self.ckpt = CheckpointManager(output_dir, input_path=input_path)
        # ★ 4.0.27: 兼容舊 APIMonitor (客戶端 checkpoint_manager.py 不在 whitelist 永遠是 4.0.10 版,
        #   不認 event_log_fn 參數). 用 try/except 試新 signature, 失敗退回舊.
        try:
            self.monitor = APIMonitor(fail_threshold=8, success_reset=5, event_log_fn=self.log_detail)
        except TypeError:
            self.monitor = APIMonitor(fail_threshold=8, success_reset=5)
        set_monitor(self.monitor)
        # ★ Resume 條件: 有 ckpt + 有 xlsx + 有 stage_progress
        # 注意: 不能要求 completed_stages 非空 — V64 中段 PauseException 時 completed_stages 是 []
        # 但 stage_progress 有 (Stage A/B/C/D 內部已 done), 必須 resume 才能跳過已 done items
        self._resume_mode = (resume and CheckpointManager.has_checkpoint(output_dir)
                             and os.path.exists(self.ckpt.xlsx_path)
                             and bool(self.ckpt.state.get('stage_progress')))
        # ★ Bug B 修: Resume 時驗證 input 一致 (用戶換了 input 就警告)
        if self._resume_mode:
            saved_input = self.ckpt.state.get('input_path', '')
            if saved_input and saved_input != input_path:
                # 比較 basename, 給 warning 但不阻擋 (用戶可能挪過位置)
                from os.path import basename
                if basename(saved_input) != basename(input_path):
                    self.log(f"\n⚠️ [Resume 警告] input 文件不一致!")
                    self.log(f"  ckpt 記錄: {saved_input}")
                    self.log(f"  本次輸入: {input_path}")
                    self.log(f"  → 為避免結果錯亂, 退回新跑模式 (將從頭處理本次 input)")
                    self._resume_mode = False
        if self._resume_mode:
            done_stages = ', '.join(self.ckpt.state['completed_stages']) or '(無 stage 整體完成)'
            in_progress = list(self.ckpt.state.get('stage_progress', {}).keys())
            self.log(f"\n[★ Resume] 從 checkpoint 接續")
            self.log(f"  已完成 stages: {done_stages}")
            self.log(f"  進行中 stages: {', '.join(in_progress) if in_progress else '(無)'}")
            self.ckpt.clear_paused()  # 清掉 paused flag, 開跑
        else:
            # 全新跑 → 重置 ckpt (避免舊 ckpt 殘留)
            self.ckpt = CheckpointManager(output_dir, input_path=input_path)

        t0 = time.time()
        self.log(f"{'='*50}")
        self.log(f"開始處理 | 來源: {source_type} | 文件: {os.path.basename(input_path)}")
        self.log(f"{'='*50}")

        # ★ Feedback collector: 跑批時收 LaMa events, 結束時 push 到 admin
        # Best-effort, 失敗永遠不 block 用戶跑批
        try:
            from processors.feedback_collector import get_collector
            tg_id = str(self.config.get('tg_id', '')).strip()
            self._feedback_collector = get_collector()
            self._feedback_collector.configure(tg_id=tg_id)
            batch_id = self._feedback_collector.get_batch_id()
            self._fp_batch_id = batch_id
            self.log(f"  [feedback] batch_id={batch_id} (給 admin Claude review 用)")
        except Exception as e:
            self.log(f"  [feedback] init 失敗: {e}")
            self._feedback_collector = None

        # ★ 4.0.25: ADMIN_FINGERPRINT dump (batch_id 已就緒, input_rows 在 Step 0 後加進來)
        try:
            fp = self._collect_admin_fingerprint()
            self.log_admin('FINGERPRINT', fp)
            # ★ 4.0.32: 用 fingerprint 測到的 mw_ping_ms 設 hub 端並發 cap
            #   高 RTT (218ms 同事踩過) 下中介給的 76 並發 → TLS 握手撐爆
            #   hub 端自己 cap, 不靠中介改
            try:
                from processors.utils import set_hub_rtt, get_hub_cap_info
                rtt_raw = fp.get('mw_ping_ms')
                if isinstance(rtt_raw, int):
                    set_hub_rtt(rtt_raw)
                    cap_info = get_hub_cap_info()
                    if 'cap=' in cap_info:
                        self.log(f"[網路] mw_ping={rtt_raw}ms → 並發 {cap_info} (高 RTT 下避免 TLS 撐爆)")
            except Exception:
                pass
        except Exception:
            pass

        # ─── Step 0: 讀取並適配 ───
        # ★ Resume 模式: 直接從 _checkpoint.xlsx 載入 df, 跳過 adapt + Step A-G
        if self._resume_mode and os.path.exists(self.ckpt.xlsx_path):
            self.log(f"\n[Step 0] ⏭ Resume 從 _checkpoint.xlsx 載入 df...")
            try:
                df = pd.read_excel(self.ckpt.xlsx_path, engine='openpyxl')
                # ★ Excel I/O NaN bug — resume 致命:
                #   pandas 讀 Excel 時把空字串 → NaN; 但下游 (seo_v64_full / title_dedup / pipeline)
                #   大量用 `str(df.at[idx, col] or default)` pattern 期待空字串 fallback.
                #   而 `bool(NaN) == True` 所以 `NaN or default` 不 fallback, str(NaN) = 'nan' (字串).
                #   後果: forbiddens_text 全變 'nan' → 標違禁 → 全部商品不合格 (4.0.11→4.0.12 修).
                #
                #   修法範圍 (audit 完整覆蓋, 4.0.13):
                #     A. _v65_*           — Stage A-E 內部 cache 欄 (forbidden/detail/subtitle/seo/...)
                #     B. _filter_*        — 過濾原因 / 標記
                #     C. 用戶面文字欄     — LLM 輸入用 (NaN 標題會被當 'nan' 字串送 LLM)
                #
                #   不修法選擇 (deliberate):
                #     × 數值欄 (起標價/數量/淘宝分類ID) — 不該強轉 str, 後面會被 numeric 比較
                #     × 商品條碼 — 已在各處 str() cast, 不依賴 fillna
                _STR_COLS_NEED_NAN_FIX = {
                    '標題', '說明', '商品簡述', '標籤', '圖片',
                    '拍賣類別名稱', '淘宝分類名稱',
                }
                _fixed_cols = []
                for col in df.columns:
                    if (col.startswith('_v65_') or col.startswith('_filter_')
                            or col in _STR_COLS_NEED_NAN_FIX):
                        s = df[col].fillna('').astype(str)
                        # mask 殘留 'nan' 字串 (舊 ckpt 已被污染, 或 dtype=float64 整列 NaN→astype 變 'nan')
                        df[col] = s.mask(s == 'nan', '')
                        _fixed_cols.append(col)
                self.log(f"  載入: {len(df)} 行 | 已完成 stages: {', '.join(self.ckpt.state['completed_stages'])}")
                self.stats['input_rows'] = len(df)
                self._fp_input_rows = len(df)  # 4.0.29: resume mode 也設, ADMIN_FINGERPRINT 才不會印 0
            except Exception as e:
                # ★ Bug C 修: ckpt.xlsx 損壞時 fallback 新跑模式
                self.log(f"\n⚠️ [Resume 失敗] _checkpoint.xlsx 讀取失敗: {e}")
                self.log(f"  → 退回新跑模式, 從原 input 重新處理")
                self._resume_mode = False
                # 刪掉壞 ckpt 避免下次又踩
                try:
                    os.remove(self.ckpt.xlsx_path)
                    os.remove(self.ckpt.json_path)
                except Exception: pass
                self.ckpt = CheckpointManager(output_dir, input_path=input_path)
                # 進入新跑流程 (下面 else 分支)
                df = None
        else:
            df = None
        # ★ force_raw 一律定義 (resume 模式下也需要, Step H 等會用到)
        process_mode = self.config.get('process_mode', 'raw')
        force_raw = (process_mode == 'raw')
        if not self._resume_mode:
            self.log(f"\n[Step 0] 讀取輸入文件... (模式: {'原始數據' if process_mode == 'raw' else '已處理數據'})")
            raw_df = pd.read_excel(input_path, engine='openpyxl')
            self.log(f"原始數據: {len(raw_df)} 行, {len(raw_df.columns)} 列")
            self._fp_input_rows = len(raw_df)  # 4.0.25: 給 ADMIN_SUMMARY 用

            # ★ 配額檢查 (跑貴功能前) — Stage E / 去水印 / 首圖優化
            steps_cfg = self.config.get('steps', {})
            has_expensive = QuotaClient.has_expensive_step(steps_cfg)
            qc = QuotaClient(self.config)
            self._quota_client = qc
            self._quota_consumed = 0

            # ★ 4.0.16: 配額不足 (TG ID 空 / limit=0 / remaining=0) → 自動降級, 不再 raise.
            #   關掉 3 個貴功能, 繼續跑基本流程 (分類/價格/關鍵詞/V65 Stage A-D/標題去重等).
            #   理由: 用戶請求「配額沒了 應該只取消貴功能 不該整批不能跑」.
            #   防禦在 app.py 那層已先取消勾選, 這裡是 defense in depth (萬一手動改 config / 別 UI bypass).
            def _disable_expensive_in_steps(reason: str):
                disabled = []
                names = {
                    'v65_stage_e': 'Stage E',
                    'image_dewatermark': '去水印救援',
                    'image_opt': '首圖 AI 優化',
                }
                for key in QuotaClient.EXPENSIVE_STEPS:
                    cur = steps_cfg.get(key, {})
                    if cur.get('enabled'):
                        cur['enabled'] = False
                        steps_cfg[key] = cur
                        disabled.append(names.get(key, key))
                if disabled:
                    self.log(f"\n⚠️ [配額] {reason}")
                    self.log(f"   → 自動關閉貴功能: {', '.join(disabled)}")
                    self.log(f"   → 繼續跑基本流程 (分類/價格/關鍵詞/V65 Stage A-D/標題去重 等), 不需配額")

            if has_expensive and qc.is_active():
                if not qc.tg_id:
                    _disable_expensive_in_steps("未填 TG ID, 跑貴功能必須填")
                    has_expensive = False
                else:
                    # 查當前配額
                    try:
                        status = qc.precheck(len(raw_df), has_expensive=True)
                    except QuotaConfigError as e:
                        self.log(f"\n⚠️ [配額] 配置不完整: {e}")
                        _disable_expensive_in_steps("配額後端配置不完整")
                        has_expensive = False
                        status = None
                    except QuotaNetworkError as e:
                        # 網路問題仍 raise — 因為不知道實際剩多少, 不該瞎跑貴功能扣壞配額
                        self.log(f"\n❌ [配額] 連線失敗 ({e}). 中止處理避免誤扣費.")
                        raise

                    if status is not None:
                        used = status.get('used', 0)
                        limit = status.get('limit', 0)
                        remaining = status.get('remaining', 0)
                        next_reset = status.get('next_reset_utc', '')
                        self.log(f"[配額] 今日 已用 {used}/{limit}, 剩 {remaining} 件 (下次重置: {next_reset})")

                        # ★ 額度為 0 → 不再 raise, 改自動關貴功能繼續跑
                        if limit == 0:
                            _disable_expensive_in_steps(f"額度為 0 (admin 還沒給額度)")
                            has_expensive = False
                        elif remaining == 0:
                            _disable_expensive_in_steps(
                                f"今日已用滿 ({used}/{limit}). 等 10:00 (台北時間) 自動重置, 或請管理員加額"
                            )
                            has_expensive = False
                        else:
                            # ★ 仍有配額 — 跑超量拆檔 + 預扣
                            # 超量 → 自動拆檔 (處理前 remaining 件, 剩餘存 overflow)
                            total_in = len(raw_df)
                            if total_in > remaining:
                                overflow_n = total_in - remaining
                                in_dir = os.path.dirname(input_path)
                                # ★ 智慧檔名: 砍掉舊的 _overflow_N件 後綴 (含舊版疊加產生的多層), 避免檔名爆長
                                #   例 1: "X_overflow_6636件.xlsx" → 拆後存 "X_overflow_4636件.xlsx"
                                #   例 2: "X_overflow_6636件_overflow_4636件.xlsx" (舊版產的) → 也乾淨剝成 "X"
                                in_stem = re.sub(r'(_overflow_\d+件)+$', '', Path(input_path).stem)
                                overflow_path = os.path.join(in_dir, f"{in_stem}_overflow_{overflow_n}件.xlsx")
                                # ★ 碰撞保護: 同名檔已存在 → 加時間戳避免覆寫
                                if os.path.exists(overflow_path):
                                    from datetime import datetime as _dt
                                    ts = _dt.now().strftime('%H%M%S')
                                    overflow_path = os.path.join(in_dir, f"{in_stem}_overflow_{overflow_n}件_{ts}.xlsx")
                                try:
                                    raw_df.iloc[remaining:].to_excel(overflow_path, index=False, engine='openpyxl')
                                    raw_df = raw_df.iloc[:remaining].copy()
                                    self.log(f"\n⚠️ [配額] 超量 {total_in} > {remaining}, 已自動拆檔:")
                                    self.log(f"   本批處理: {remaining} 件")
                                    self.log(f"   剩餘 {overflow_n} 件 存到: {overflow_path}")
                                    self.log(f"   ↳ 等明日 10:00 重置後, 再用此 overflow 檔重跑")
                                except Exception as e:
                                    self.log(f"\n❌ [配額] 拆檔失敗: {e}")
                                    raise

                            # 預扣 (原子操作 — 多機併發安全)
                            consume_n = len(raw_df)
                            try:
                                after = qc.consume(consume_n)
                                self._quota_consumed = consume_n
                                self.log(f"[配額] 已扣 {consume_n} 件 (剩 {after.get('remaining', '?')}/{after.get('limit', '?')})")
                            except QuotaExceeded as e:
                                # 並發競爭: 其他機器搶先扣完了 → 自動降級, 不 raise
                                self.log(f"\n⚠️ [配額] 預扣失敗 (其他機器併發搶先?): {e}")
                                _disable_expensive_in_steps("並發搶先, 配額剛被別人扣光")
                                has_expensive = False

            if source_type == 'mercari':
                df = adapt_mercari(raw_df, force_raw=force_raw, log_fn=self.log)
            else:
                df = adapt_goofish(raw_df, force_raw=force_raw, log_fn=self.log)

            self.stats['input_rows'] = len(df)
        self._check_stop()

        # ★ 4.0.33: deep copy steps snapshot — 跑批中 GUI thread (_update_expensive_lockout)
        # 看到 quota=0 不該改 config 害 pipeline 中途 Step K0/K 看到 enabled=False 跳過.
        # GUI 端已加 guard (跑批中 early-return), pipeline 端 deep copy 是 defense in depth.
        # 配額 graceful degrade (_disable_expensive_in_steps) 在 line 333 之前已跑完, deep copy
        # 拿到的是已 graceful degrade 後的版本, 後續 stage 仍正確跳過.
        import copy as _copy
        steps = _copy.deepcopy(self.config.get('steps', {}))

        # ★ Resume: Step A-G 全跳過 (已包含在 _checkpoint.xlsx)
        if self._resume_mode:
            self.log(f"\n[Resume] ⏭ 跳過 Step A-G (df 已從 _checkpoint.xlsx 載入)")

        # ─── Step A: 分類映射 (最先過濾標紅, 省API) ───
        if not self._resume_mode and steps.get('category', {}).get('enabled'):
            _progress('分類')
            self.log("\n[Step A] 分類處理...")
            self._run_category(df, source_type)
            self._check_stop()

        # ─── Step B: 價格轉換 (過濾過高過低, 省API) ───
        if not self._resume_mode and steps.get('price', {}).get('enabled'):
            _progress('價格')
            self.log("\n[Step B] 價格轉換...")
            self.stats['price'] = apply_price_conversion(
                df, source_type, self.config.get('price', {}), self.log)
            self._check_stop()

        # ─── Step C: 翻譯 (煤爐日→繁, 只處理合格商品) ───
        if not self._resume_mode and steps.get('translate', {}).get('enabled'):
            _progress('翻譯')
            has_reason = df['_filter_reason'].fillna('').str.strip().astype(bool)
            skip_count = has_reason.sum()
            if skip_count > 0:
                self.log(f"\n[Step C] 翻譯處理... (跳過{skip_count}條已過濾商品)")
            else:
                self.log("\n[Step C] 翻譯處理...")

            if (~has_reason).any():
                good_df = df[~has_reason].copy()
                self._run_translate(good_df, source_type)
                for col in ['標題', '說明']:
                    if col in good_df.columns:
                        df.loc[good_df.index, col] = good_df[col]
            else:
                self.log("  所有商品已被過濾, 跳過翻譯")
            self._check_stop()

        # ─── Step D: 替換詞 ───
        if not self._resume_mode and steps.get('replace', {}).get('enabled'):
            _progress('替換詞')
            self.log("\n[Step D] 替換詞處理...")
            try:
                pairs = load_replacements(self.config['paths']['replace_xlsx'], self.log)
                self.stats['replace'] = apply_replacements(df, pairs, self.log)
            except Exception as e:
                self.log(f"[錯誤] 替換詞處理失敗: {e}")
            self._check_stop()

        # ─── Step E: 價格/數字清洗 (標題+說明) ───
        if not self._resume_mode and steps.get('price_clean', {}).get('enabled'):
            _progress('價格清洗')
            self.log("\n[Step E] 價格/數字清洗...")
            try:
                paths = self.config.get('paths', {})
                clean_rules = load_price_cleaner(
                    paths.get('number_removal_xlsx', ''),
                    paths.get('full_removal_xlsx', ''),
                    paths.get('protect_numbers_xlsx', ''),
                    self.log)
                self.stats['price_clean'] = apply_price_cleaning(df, clean_rules, self.log)
            except Exception as e:
                self.log(f"[錯誤] 價格清洗失敗: {e}")
            self._check_stop()

        # ─── Step G: 關鍵詞違禁過濾 (rule-based, 提前到 LLM 之前省 token) ───
        if not self._resume_mode and steps.get('keyword', {}).get('enabled'):
            _progress('關鍵詞')
            self.log("\n[Step G] 關鍵詞過濾 (LLM 前)...")
            try:
                kw_list = load_keywords(self.config['paths']['keyword_xlsx'], self.log)
                self.stats['keyword'] = apply_keyword_filter(df, kw_list, self.log)
            except Exception as e:
                self.log(f"[錯誤] 關鍵詞過濾失敗: {e}")
            self._check_stop()

        # ─── Step F-V65: ★ 多模態 SEO + 視覺違禁 (取代 F + F2) ───
        # Stage A-D 一定跑 (互相依賴), Stage E 可獨立關閉
        v64_enabled = steps.get('seo_v64_image', {}).get('enabled', False)
        stage_e_enabled = steps.get('v65_stage_e', {}).get('enabled', True)  # 預設開
        if v64_enabled:
            # ★ Resume: V64 已完成則跳過 (df 已從 ckpt xlsx 載入)
            if self._resume_mode and self.ckpt.is_stage_done('seo_v64'):
                self.log(f"\n[Step F-V65] ⏭ Resume 跳過 (上次已完成)")
            else:
                _progress('V64 多模態 SEO+視覺違禁')
                self.log(f"\n[Step F-V65] 多模態 SEO + 視覺違禁判定 (Stage A-D{' + Stage E' if stage_e_enabled else ', Stage E 關'})...")
                seo_dir = self.config['paths'].get('seo_tool_dir', '')
                self._expensive_started = True  # ★ 貴功能開跑 — 之後 abort 不退費
                self.ckpt.start_stage('seo_v64', df=df)
                try:
                    self.stats['seo_v64'] = run_v64_full(df, self.log, seo_dir, enable_stage_e=stage_e_enabled, log_detail_fn=self.log_detail, ckpt=self.ckpt)
                    self.ckpt.complete_stage('seo_v64', df=df)
                except PauseException as pe:
                    return self._handle_pause(pe, df, t0)
                except Exception as e:
                    self.log(f"[警告] V64 失敗: {e}")
                    import traceback; traceback.print_exc()
            self._check_stop()

        # ─── Step F: SEO標題優化 (需API, 只處理合格商品) ───
        if not v64_enabled and steps.get('seo', {}).get('enabled'):
            _progress('SEO')
            # 只對尚未被過濾的商品做SEO (省API)
            has_reason = df['_filter_reason'].fillna('').str.strip().astype(bool)
            skip_count = has_reason.sum()
            if skip_count > 0:
                self.log(f"\n[Step F] SEO標題優化... (跳過{skip_count}條已過濾商品)")
            else:
                self.log("\n[Step F] SEO標題優化...")

            if (~has_reason).any():
                good_df = df[~has_reason].copy()
                self._run_seo(good_df, source_type)
                # 回寫結果到原df (Step F 只處理標題+說明)
                for col in ['標題', '說明']:
                    if col in good_df.columns:
                        if col not in df.columns:
                            df[col] = ''
                        df.loc[good_df.index, col] = good_df[col]
                if '_filter_reason' in good_df.columns:
                    df.loc[good_df.index, '_filter_reason'] = good_df['_filter_reason']
            else:
                self.log("  所有商品已被過濾, 跳過SEO")
            self._check_stop()

        # ─── Step F2: 商品簡述 + 5 標籤生成 (V44/V42, 獨立步驟, 需API) ───
        if not v64_enabled and steps.get('f2_subtitle_hashtag', {}).get('enabled'):
            _progress('簡述標籤')
            has_reason = df['_filter_reason'].fillna('').str.strip().astype(bool)
            skip_count = has_reason.sum()
            if skip_count > 0:
                self.log(f"\n[Step F2] 商品簡述 + 標籤生成... (跳過{skip_count}條已過濾商品)")
            else:
                self.log("\n[Step F2] 商品簡述 + 標籤生成...")

            if (~has_reason).any():
                good_df = df[~has_reason].copy()
                seo_dir = self.config['paths'].get('seo_tool_dir', '')
                try:
                    self._run_subtitle_hashtag(good_df, seo_dir)
                    # V45: 回寫 簡述 + 標籤 + 清洗後說明 + _filter_reason (違禁品攔截)
                    for col in ['商品簡述', '標籤', '說明', '_filter_reason']:
                        if col in good_df.columns:
                            if col not in df.columns:
                                df[col] = ''
                            df.loc[good_df.index, col] = good_df[col]
                except Exception as e:
                    self.log(f"[警告] F2 簡述/標籤/說明生成失敗: {e}")
            else:
                self.log("  所有商品已被過濾, 跳過 F2")
            self._check_stop()

        # ─── Step H: 默認值填充 ───
        if steps.get('defaults', {}).get('enabled'):
            _progress('默認值')
            self.log("\n[Step H] 默認值填充...")
            defs = self.config.get('defaults', {}).get(source_type, {})
            self.stats['defaults'] = apply_defaults(
                df, source_type, defs, force_overwrite=force_raw, log_fn=self.log)
            self._check_stop()

        # ─── Step H: AI過濾 (可選, 需API) ───
        if steps.get('ai_filter', {}).get('enabled'):
            _progress('AI過濾')
            self.log("\n[Step H] AI商品過濾...")
            self._run_ai_filter(df)
            self._check_stop()

        # ─── Step I: 標題去重 ───
        if steps.get('dedup', {}).get('enabled'):
            _progress('去重')
            self.log("\n[Step I] 標題去重...")
            self.stats['dedup'] = apply_dedup(df, self.log)
            self._check_stop()

        # ─── Step J: 說明模板追加 ───
        if steps.get('desc_append', {}).get('enabled'):
            _progress('說明模板')
            self.log("\n[Step J] 說明模板追加...")
            templates_cfg = self.config.get('desc_templates', {})
            selected_name = templates_cfg.get('selected', '')
            tpl_list = templates_cfg.get('templates', [])
            tpl = next((t for t in tpl_list if t['name'] == selected_name), None)
            if tpl:
                self.stats['desc_append'] = apply_desc_append(
                    df, tpl['content'], tpl.get('color', ''),
                    tpl.get('font_family', ''), tpl.get('font_size', ''),
                    tpl.get('position', 'before'), self.log)
            else:
                self.log(f"[警告] 找不到所選模板 '{selected_name}', 跳過說明模板")
            self._check_stop()

        # ─── Step K0: 去水印救援 (AI 定位 + LaMa, 對 V65 reject 圖) ───
        # V65 Stage E 標記 reject 的圖, 嘗試 LaMa 救援保留 (商品 100% 保真)
        # 救援失敗 (面積>30% 或廣告整圖) 才仍丟棄
        if v64_enabled and steps.get('image_dewatermark', {}).get('enabled', True):
            if self._resume_mode and self.ckpt.is_stage_done('k0_dewatermark'):
                self.log(f"\n[Step K0] ⏭ Resume 跳過 (上次已完成)")
            else:
                _progress('去水印救援')
                self.log("\n[Step K0] 去水印救援 (AI+LaMa)...")
                self._expensive_started = True  # ★ 貴功能 — 後續 abort 不退費
                self.ckpt.start_stage('k0_dewatermark', df=df)
                try:
                    lama_cfg = self.config.get('lama', {}) or {}
                    self.stats['image_dewater'] = try_dewatermark_rejected(
                        df, self.log,
                        log_detail_fn=self.log_detail,
                        ckpt=self.ckpt,
                        lama_device=lama_cfg.get('device', 'auto'),
                        lama_pool_size=lama_cfg.get('pool_size'),
                    )
                    self.ckpt.complete_stage('k0_dewatermark', df=df)
                except PauseException as pe:
                    return self._handle_pause(pe, df, t0)
                except Exception as e:
                    self.log(f"[警告] 去水印救援失敗: {e}")
                    import traceback; traceback.print_exc()
            self._check_stop()

        # ─── Step K: 首圖 AI 優化 (GPT-image-2, 在輸出前) ───
        # 把每商品圖片資料夾的第 1 張圖跑 GPT-image-2 優化, 結果存為 0.jpg, 排在原圖前
        # 違禁/低分商品自動跳過 (省 image API 配額)
        if steps.get('image_opt', {}).get('enabled'):
            if self._resume_mode and self.ckpt.is_stage_done('image_opt'):
                self.log(f"\n[Step K] ⏭ Resume 跳過 (上次已完成)")
            else:
                _progress('首圖優化')
                self.log("\n[Step K] 首圖 AI 優化 (GPT-image-2)...")
                self._expensive_started = True  # ★ 貴功能 — 後續 abort 不退費
                self.ckpt.start_stage('image_opt', df=df)
                try:
                    self.stats['image_opt'] = optimize_first_images(df, self.log, log_detail_fn=self.log_detail, ckpt=self.ckpt)
                    self.ckpt.complete_stage('image_opt', df=df)
                except PauseException as pe:
                    return self._handle_pause(pe, df, t0)
                except Exception as e:
                    self.log(f"[警告] 首圖優化失敗: {e}")
            self._check_stop()

        # ─── 精準簡繁轉換 (只轉 Yahoo 真實黑名單字, 保留其他繁體寫法) ───
        # 對齊 Yahoo bundle 反編譯黑名單 (4563 字), 不動「家具/托盤/回流」等慣用詞
        try:
            from opencc import OpenCC
            blacklist_path = os.path.join(os.path.dirname(__file__), 'yahoo_hans_blacklist.txt')
            blacklist = set()
            if os.path.exists(blacklist_path):
                with open(blacklist_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        if '\t' in line:
                            ch = line.split('\t')[0]
                            if ch: blacklist.add(ch)
            if not blacklist:
                self.log(f"\n[警告] 黑名單載入失敗, 跳過簡繁轉換")
            else:
                cc = OpenCC('s2tw')
                # ── 詞級覆寫 (專有名詞/詞組, 比字級早跑) ──
                WORD_OVERRIDE = {
                    '鎌倉雕': '日式漆雕',    # 日本 Kamakura-bori
                    '鎌倉': '日式古典',      # 鎌倉時代/地名 (萬一非雕)
                    '青銅鐓': '青銅戈鐏',    # 戈柄末端銅帽 (故宮級用詞)
                    '軲轆': '轆轤',          # 圓物/井絞盤
                    '鋥亮': '光亮',          # 拋光鋥亮 → 拋光光亮
                    '柺杖': '拐杖',          # Taiwan 標準
                }
                # ── 字級覆寫 (OpenCC 救不了的字 fallback) ──
                CHAR_OVERRIDE = {
                    '獁': '瑪',  # 猛獁 → 猛瑪 (mammoth)
                    '叁': '三',  # 簡體大寫三 → 普通三 (Taiwan 銀貨/紙幣讀法)
                    '叄': '三',  # 異體大寫三
                    '鎌': '鐮',  # 鐮刀 (詞級漏網的單獨 鎌)
                    '鐓': '鐏',  # 戈柄末端銅帽 (詞級漏網)
                    '鋥': '光',  # 鋥光 → 光
                    '軲': '轂',  # 詞級漏網
                    '柺': '拐',  # 詞級漏網
                    '擼': '抓',  # 大陸口語
                }
                # 預先建構字典: CHAR_OVERRIDE 優先, 否則用 cc.convert
                _blacklist_map = {}
                _unresolved_set = set()  # set 而非 list, in O(1)
                for ch in blacklist:
                    if ch in CHAR_OVERRIDE:
                        _blacklist_map[ch] = CHAR_OVERRIDE[ch]
                    else:
                        conv = cc.convert(ch)
                        _blacklist_map[ch] = conv
                        if len(conv) == 1 and conv in blacklist:
                            _unresolved_set.add(ch)

                # 詞級替換按 key 長度遞減 (長詞優先, 避免短詞先吃掉)
                _word_pairs = sorted(WORD_OVERRIDE.items(), key=lambda x: -len(x[0]))

                def smart_convert(text):
                    if not text: return text
                    # Step 1: 詞級替換
                    for old, new in _word_pairs:
                        if old in text:
                            text = text.replace(old, new)
                    # Step 2: 字級替換
                    return ''.join(_blacklist_map.get(c, c) for c in text)

                n_changed = 0
                n_chars_changed = 0
                # fallback 統計: 仍踩雷的字 (CHAR_OVERRIDE/OpenCC 都救不了的)
                fallback_hits = {}  # {char: [商品條碼...]}
                for col in ['標題', '商品簡述', '說明', '標籤']:
                    if col not in df.columns: continue
                    for idx in df.index:
                        v = df.at[idx, col]
                        if isinstance(v, str) and v:
                            new_v = smart_convert(v)
                            if new_v != v:
                                df.at[idx, col] = new_v
                                n_changed += 1
                                n_chars_changed += sum(1 for a, b in zip(v, new_v) if a != b)
                            # 二次掃描: 若仍有黑名單字, fallback 跳過 (刪除) + 記錄
                            still_bad = [c for c in new_v if c in blacklist]
                            if still_bad:
                                code = str(df.at[idx, '商品條碼']) if '商品條碼' in df.columns else f'idx={idx}'
                                for c in still_bad:
                                    fallback_hits.setdefault(c, []).append(code)
                                # 跳過策略: 刪除這些字, 商品仍能 publish
                                df.at[idx, col] = ''.join(c for c in new_v if c not in blacklist)
                self.log(f"\n[簡繁轉換] 完成 (黑名單 {len(blacklist)} 字), 改 {n_changed} 個欄位 / {n_chars_changed} 個字")
                if fallback_hits:
                    self.log(f"[警告] {len(fallback_hits)} 個字 OVERRIDE 沒覆蓋, 已自動刪除以保 publish:")
                    for c, codes in sorted(fallback_hits.items(), key=lambda x: -len(x[1])):
                        self.log(f"        '{c}' (U+{ord(c):04X}) × {len(codes)} 件 — 商品: {','.join(codes[:3])}{'...' if len(codes)>3 else ''}")
                    # 持久化寫檔: output_dir/_unknown_chars_YYYYMMDD.txt
                    try:
                        from datetime import datetime
                        report_dir = self.config.get('output_dir', '') or os.path.dirname(input_path)
                        os.makedirs(report_dir, exist_ok=True)
                        report_path = os.path.join(report_dir, f"_unknown_chars_{datetime.now().strftime('%Y%m%d')}.txt")
                        with open(report_path, 'a', encoding='utf-8') as rf:
                            rf.write(f"\n{'='*60}\n")
                            rf.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 來源: {os.path.basename(input_path)}\n")
                            rf.write(f"本次發現 {len(fallback_hits)} 個 OVERRIDE 沒覆蓋的字 (已自動刪除):\n\n")
                            for c, codes in sorted(fallback_hits.items(), key=lambda x: -len(x[1])):
                                rf.write(f"  '{c}' (U+{ord(c):04X}) × {len(codes)} 件\n")
                                rf.write(f"    商品: {', '.join(codes)}\n\n")
                            rf.write(f"建議: 補進 pipeline.py:469 的 CHAR_OVERRIDE 或 WORD_OVERRIDE\n")
                        self.log(f"[警告] 詳細記錄已寫入: {report_path}")
                    except Exception as e:
                        self.log(f"[警告] 寫入 _unknown_chars 報告失敗: {e}")
        except ImportError:
            self.log(f"\n[警告] opencc 未安裝, 跳過簡繁轉換")
        except Exception as e:
            self.log(f"\n[警告] 簡繁轉換失敗: {e}")

        # ─── 標籤硬上限 5 個 (Yahoo 規則, 多餘的截掉) ───
        if '標籤' in df.columns:
            n_truncated = 0
            for idx in df.index:
                v = df.at[idx, '標籤']
                if not isinstance(v, str) or not v: continue
                tags = v.split()
                if len(tags) > 5:
                    df.at[idx, '標籤'] = ' '.join(tags[:5])
                    n_truncated += 1
            if n_truncated > 0:
                self.log(f"\n[標籤上限] 截斷 {n_truncated} 件商品的標籤至 5 個 (Yahoo 上限)")

        # ─── 輸出拆分 ───
        self.log(f"\n{'='*50}")
        self.log("拆分輸出...")
        result = self._split_and_export(df, input_path, source_type)

        elapsed = time.time() - t0
        self.log(f"\n處理完成! 耗時 {elapsed:.1f}s")
        self.log(f"  合格: {result['good_count']}")
        self.log(f"  不合格: {result['bad_count']}")
        self.log(f"  未分類: {result['uncat_count']}")

        # ★ 不合格分組總結 (詳細日誌)
        if '_filter_reason' in df.columns and result['bad_count'] > 0:
            from collections import Counter
            reasons = df[df['_filter_reason'].fillna('').str.strip().astype(bool)]['_filter_reason'].tolist()
            # 簡化原因 (第一個分號前)
            simplified = [str(r).split(';')[0].split(':')[0].strip() for r in reasons]
            counter = Counter(simplified)
            self.log(f"  不合格分組:")
            for reason, n in counter.most_common(10):
                self.log(f"    {reason}: {n}")
            # 詳細日誌列出每件商品
            self.log_detail(f"\n[拆分輸出] 不合格商品詳情 ({result['bad_count']} 件):")
            for idx, row in df.iterrows():
                fr = str(row.get('_filter_reason') or '').strip()
                if fr:
                    code = str(row.get('商品條碼', ''))[:20]
                    self.log_detail(f"    ✗ {code} | {fr[:100]}")

        result['stats'] = self.stats
        result['elapsed'] = elapsed
        result['paused'] = False
        # ★ 全部跑完: 清掉 checkpoint
        if self.ckpt:
            self.ckpt.complete_stage('final_export', df=df)
            self.ckpt.cleanup()
            self.log("[Checkpoint] 已清除 (全部完成)")
        # ★ Push feedback (LaMa events) 給 admin (best-effort, 不 block)
        try:
            if hasattr(self, '_feedback_collector') and self._feedback_collector and self._feedback_collector.event_count() > 0:
                n = self._feedback_collector.event_count()
                self.log(f"[feedback] push {n} 個 LaMa events 給 admin...")
                self._feedback_collector.push(log_fn=lambda m, level='info': self.log(m))
        except Exception as e:
            self.log(f"[feedback] push 失敗 (不影響結果): {e}")
        # ★ 4.0.25: admin summary (給 admin Claude 從詳細日誌看狀態)
        self._dump_admin_summary(
            'completed',
            good=result.get('good_count', 0),
            bad=result.get('bad_count', 0),
            uncategorized=result.get('uncat_count', 0),
        )
        # ★ 4.0.43: 關檔 + push 詳細日誌 (background thread, 不阻塞 GUI)
        # ★ 4.0.44: 改用 _finalize_detail_log idempotent + background, finally 兜底會跑.
        self._finalize_detail_log()
        return result

    def _handle_pause(self, pe: 'PauseException', df, t0: float) -> dict:
        """暫停: 寫 checkpoint + 回 paused result, GUI 顯示「等 API 恢復後按繼續」"""
        reason = str(pe)
        self.log(f"\n{'='*50}")
        self.log(f"[⏸ 暫停] {reason}")
        self.log(f"  進度已寫入 _checkpoint.* (output_dir)")
        self.log(f"  ★ 不要關閉軟件, 等 API 恢復後按「▶ 繼續」")
        self.log(f"{'='*50}")
        self._paused_during_run = True
        if self.ckpt:
            self.ckpt.mark_paused(reason, df=df)
        # ★ Pause 時也 push 已收到的 feedback (best-effort, 不 block 用戶)
        try:
            if hasattr(self, '_feedback_collector') and self._feedback_collector and self._feedback_collector.event_count() > 0:
                n = self._feedback_collector.event_count()
                self.log(f"[feedback] pause 時 push {n} 個 events")
                self._feedback_collector.push(log_fn=lambda m, level='info': self.log(m))
        except Exception:
            pass
        elapsed = time.time() - t0
        # ★ 4.0.25: admin summary on pause
        self._dump_admin_summary('paused', pause_reason=reason)
        # ★ 4.0.43/44: pause 時也 push 詳細日誌 (debug 暫停原因, admin 用)
        self._finalize_detail_log()
        return {
            'paused': True,
            'paused_reason': reason,
            'elapsed': elapsed,
            'stats': self.stats,
            'good_count': 0, 'bad_count': 0, 'uncat_count': 0,
            'good_path': '', 'bad_path': '', 'uncat_path': '',
        }

    # ─── 翻譯橋接 ───
    def _run_translate(self, df, source_type):
        """調用 SEO翻譯工具 的翻譯功能"""
        if source_type != 'mercari':
            self.log("鹹魚數據已是繁中, 跳過翻譯")
            return
        try:
            seo_dir = self.config['paths'].get('seo_tool_dir', '')
            if not seo_dir or not os.path.isdir(seo_dir):
                self.log(f"[跳過] SEO翻譯工具目錄不存在: {seo_dir}")
                return
            if seo_dir not in sys.path:
                sys.path.insert(0, seo_dir)
            from translator import CFG, translate_dataframe, resolve_column_name, TITLE_SYNONYMS, DESC_SYNONYMS

            self._load_seo_config(seo_dir, CFG)
            CFG.enable_translate = True
            CFG.enable_seo = False  # SEO在下一步
            CFG.enable_kana_cleanup = True

            # 重置密鑰池
            import translator
            translator._KEY_POOL = None

            self.log(f"  翻譯配置: {CFG.api_base} | 模型: {CFG.model} | 並發: {CFG.workers} | 批次: {CFG.batch_size}")

            # 只在 5% 整數倍變動時印, 避免日誌噪音 (從每 batch 印 → 每 5% 印一次)
            _trans_last_pct = [-1]
            def _trans_progress(current, total, stage=''):
                pct = int(current / total * 100) if total else 0
                bucket = pct - (pct % 5)
                if bucket != _trans_last_pct[0]:
                    _trans_last_pct[0] = bucket
                    self.log(f"  [{stage}] {current}/{total} ({pct}%)")

            tcol = resolve_column_name(df, '標題', TITLE_SYNONYMS)
            dcol = resolve_column_name(df, '說明', DESC_SYNONYMS)
            result_df = translate_dataframe(df, tcol, dcol, log_fn=self.log,
                                            progress_fn=_trans_progress,
                                            stop_check=lambda: self._stop)
            df[tcol] = result_df[tcol]
            df[dcol] = result_df[dcol]
            self.log("翻譯完成")
        except ImportError as e:
            self.log(f"[跳過] 無法導入 SEO翻譯工具: {e}")
        except Exception as e:
            import traceback
            self.log(f"[錯誤] 翻譯失敗: {e}\n{traceback.format_exc()}")

    # ─── SEO橋接 ───
    def _run_seo(self, df, source_type):
        """調用 SEO翻譯工具 的SEO功能"""
        try:
            seo_dir = self.config['paths'].get('seo_tool_dir', '')
            if not seo_dir or not os.path.isdir(seo_dir):
                self.log(f"[跳過] SEO翻譯工具目錄不存在: {seo_dir}")
                return
            if seo_dir not in sys.path:
                sys.path.insert(0, seo_dir)
            from translator import CFG, translate_dataframe, resolve_column_name, TITLE_SYNONYMS, DESC_SYNONYMS

            self._load_seo_config(seo_dir, CFG)
            CFG.enable_translate = False
            CFG.enable_seo = True
            CFG.enable_kana_cleanup = True

            # 重置密鑰池
            import translator
            translator._KEY_POOL = None

            self.log(f"  SEO配置: {CFG.api_base} | 翻譯: {CFG.model} | SEO: {CFG.seo_model} | 並發: {CFG.workers} | 批次: {CFG.batch_size}")

            # 同樣 5% 區間印一次, 而且區分階段 (關鍵詞提取/SEO優化/重試各自獨立計算)
            _seo_last_pct = {}
            def _seo_progress(current, total, stage=''):
                pct = int(current / total * 100) if total else 0
                bucket = pct - (pct % 5)
                last = _seo_last_pct.get(stage, -1)
                if bucket != last:
                    _seo_last_pct[stage] = bucket
                    self.log(f"  [{stage}] {current}/{total} ({pct}%)")

            tcol = resolve_column_name(df, '標題', TITLE_SYNONYMS)
            dcol = resolve_column_name(df, '說明', DESC_SYNONYMS)
            result_df = translate_dataframe(df, tcol, dcol, log_fn=self.log,
                                            progress_fn=_seo_progress,
                                            stop_check=lambda: self._stop)
            df[tcol] = result_df[tcol]
            if dcol in result_df.columns:
                df[dcol] = result_df[dcol]

            # SEO評分過濾: 低於60分的歸到不合格
            if '_seo_score' in result_df.columns:
                df['_seo_score'] = result_df['_seo_score']
                low_count = 0
                for idx in df.index:
                    score = df.at[idx, '_seo_score']
                    if isinstance(score, (int, float)) and score < 60 and score > 0:
                        from processors.utils import append_reason
                        df.at[idx, '_filter_reason'] = append_reason(
                            df.at[idx, '_filter_reason'], f'SEO低分({score:.0f})')
                        low_count += 1
                if low_count:
                    self.log(f"SEO評分過濾: {low_count} 條低於60分，歸入不合格")
                df.drop(columns=['_seo_score'], inplace=True)

            self.log("SEO優化完成")

        except ImportError as e:
            self.log(f"[跳過] 無法導入 SEO翻譯工具: {e}")
        except Exception as e:
            import traceback
            self.log(f"[錯誤] SEO失敗: {e}\n{traceback.format_exc()}")

    def _run_subtitle_hashtag(self, df, seo_dir):
        """V45 Step F2: 統一 1-call 產生 商品簡述 + 5 標籤 + 清洗後說明
        (現省 65% token, 100% 成功率, 共享上下文)
        """
        if seo_dir not in sys.path:
            sys.path.insert(0, seo_dir)
        try:
            from unified_generator import gen_unified_ai
            from subtitle_generator import gen_subtitle_simple
            from hashtag_generator import gen_hashtags
            import translator
        except ImportError as e:
            self.log(f"[跳過 F2] 找不到 unified_generator: {e}")
            return

        title_col = '標題' if '標題' in df.columns else None
        brief_col = '商品簡述' if '商品簡述' in df.columns else None
        detail_col = '說明' if '說明' in df.columns else None
        if not title_col:
            self.log("[跳過 F2] 沒找到標題欄位")
            return

        if '標籤' not in df.columns:
            df['標籤'] = ''

        valid_idx = df.index[df[title_col].astype(str).str.strip().astype(bool)].tolist()
        if not valid_idx:
            self.log("[跳過 F2] 沒有有效標題")
            return

        items = []
        for idx in valid_idx:
            items.append({
                'title': str(df.at[idx, title_col] or ''),
                'brief': str(df.at[idx, brief_col] or '') if brief_col else '',
                'detail': str(df.at[idx, detail_col] or '') if detail_col else '',
            })

        self.log(f"  V45 統一 call: {len(items)} 件 → 簡述+標籤+說明清洗 (1 prompt)")
        def chat_fn(body):
            body['model'] = translator.CFG.seo_model or translator.CFG.model
            return translator._chat(body, translator.CFG.timeout)

        try:
            results = gen_unified_ai(items, chat_fn, batch_size=10, workers=20, log_fn=self.log)
        except Exception as e:
            self.log(f"  [警告] V45 統一 call 失敗, 用規則式 fallback: {e}")
            results = [{} for _ in items]

        # 寫回 df — 簡述/標籤/說明 + 違禁品攔截 (V47)
        n_sub, n_tag, n_det, n_forbid = 0, 0, 0, 0
        for i, idx in enumerate(valid_idx):
            r = results[i] if i < len(results) else {}

            # ⚠️ 違禁品攔截 (Yahoo TW 嚴禁: 煙/酒/化妝品/藥品/武器/活物)
            forbidden = (r.get('forbidden') or '').strip()
            if forbidden:
                from processors.utils import append_reason
                df.at[idx, '_filter_reason'] = append_reason(
                    df.at[idx, '_filter_reason'], f'違禁品-{forbidden}'
                )
                n_forbid += 1
                continue  # 違禁品不寫簡述/標籤/說明 (反正會被 reject)

            # 簡述: AI → 規則 fallback
            sub = (r.get('subtitle') or '').strip()
            if not sub or len(sub) < 5:
                sub = gen_subtitle_simple(items[i]['title'], items[i]['brief'], items[i]['detail'])
            if sub:
                df.at[idx, '商品簡述'] = sub
                n_sub += 1
            # 標籤: AI 為主, 接受 3-5 個 (品質優於數量), 不硬湊到 5
            tags = list(r.get('tags') or [])
            # 只在 AI 完全失敗 (<3 個) 才補, 而且至多補到 3
            if len(tags) < 3:
                supp = gen_hashtags(items[i]['title'], brief=items[i]['brief'],
                                    detail=items[i]['detail'], free_shipping=True) or []
                from unified_generator import _is_supp_valid_tag
                seen = set(tags)
                for t in supp:
                    if t not in seen and _is_supp_valid_tag(t):
                        tags.append(t)
                        seen.add(t)
                        if len(tags) >= 3: break  # 補到 3 就停, 不硬湊 5
            if tags:
                df.at[idx, '標籤'] = ' '.join(t for t in tags[:5] if t)
                n_tag += 1
            # 說明清洗
            det = (r.get('detail_clean') or '').strip()
            if det and len(det) >= 3:
                df.at[idx, '說明'] = det
                n_det += 1
            else:
                try:
                    from unified_generator import _clean_residue
                    cleaned = _clean_residue(items[i]['detail'])
                    if cleaned and len(cleaned) >= 3:
                        df.at[idx, '說明'] = cleaned
                        n_det += 1
                except Exception:
                    pass

        self.log(f"  簡述: {n_sub}/{len(valid_idx)} | 標籤: {n_tag}/{len(valid_idx)} | 說明清洗: {n_det}/{len(valid_idx)} | 違禁攔截: {n_forbid}")

    # ─── SEO配置加載 (翻譯和SEO共用) ───
    def _load_seo_config(self, seo_dir, CFG):
        """從 SEO工具的 translator_config.json 實時讀取所有配置"""
        seo_cfg_path = os.path.join(seo_dir, 'translator_config.json')
        if os.path.isfile(seo_cfg_path):
            with open(seo_cfg_path, 'r', encoding='utf-8') as f:
                seo_cfg = json.load(f)
            CFG.api_base = seo_cfg.get('api_url', CFG.api_base)
            CFG.api_key = seo_cfg.get('api_key', CFG.api_key)
            CFG.api_keys = seo_cfg.get('api_keys', [])
            CFG.model = seo_cfg.get('model', CFG.model)
            CFG.seo_model = seo_cfg.get('seo_model', CFG.seo_model)
            CFG.workers = seo_cfg.get('workers', CFG.workers)
            CFG.batch_size = seo_cfg.get('batch_size', CFG.batch_size)
            CFG.timeout = seo_cfg.get('timeout', CFG.timeout)
        else:
            self.log(f"[警告] 找不到 {seo_cfg_path}, 使用預設值")

    # ─── 分類 ───
    def _run_category(self, df, source_type):
        try:
            if source_type == 'mercari':
                mapping_path = self.config['paths']['mapping_xlsx']
                map_dict, red_ids = load_mercari_mapping(mapping_path, self.log)
                self.stats['category'] = apply_mercari_mapping(df, map_dict, red_ids, self.log)

                # 檢查未映射，嘗試自動補全
                unmapped_mask = df['_filter_reason'].str.contains('未映射', na=False)
                if unmapped_mask.any():
                    self._auto_map_unmapped(df, unmapped_mask, map_dict, red_ids)
            else:
                # 鹹魚: 優先使用映射表 (淘宝ID→奇摩ID), 無映射表時降級為規則匹配
                gf_mapping_path = self.config.get('paths', {}).get('goofish_mapping_xlsx', '')
                if gf_mapping_path and os.path.isfile(gf_mapping_path):
                    map_dict, red_ids = load_goofish_mapping(gf_mapping_path, self.log)
                    self.stats['category'] = apply_goofish_mapping(df, map_dict, red_ids, self.log)
                else:
                    rules_path = self.config['paths']['rules_xlsx']
                    rules = load_rules(rules_path, self.log)
                    self.stats['category'] = apply_rules_classification(df, rules, self.log)
        except Exception as e:
            self.log(f"[錯誤] 分類處理失敗: {e}")

    def _auto_map_unmapped(self, df, unmapped_mask, existing_map, red_ids):
        """對未映射的分類自動GPT匹配，並回填+輸出新映射"""
        paths = self.config.get('paths', {})
        full_path = paths.get('full_mercari_xlsx', '')
        yahoo_path = paths.get('yahoo_cat_xlsx', '')
        api_cfg = self.config.get('api', {})

        if not full_path or not os.path.isfile(full_path):
            self.log(f"[跳過] 煤爐完整分類表不存在: {full_path}")
            return
        if not yahoo_path or not os.path.isfile(yahoo_path):
            self.log(f"[跳過] 奇摩分類表不存在: {yahoo_path}")
            return
        if not api_cfg.get('translate_key'):
            self.log("[跳過] 未配置API Key, 無法自動映射")
            return

        # 收集未映射的分類ID
        unmapped_ids = []
        for idx in df[unmapped_mask].index:
            reason = str(df.at[idx, '_filter_reason'])
            m = re.search(r'未映射分類\((\d+)\)', reason)
            if m:
                unmapped_ids.append(int(m.group(1)))

        if not unmapped_ids:
            return

        # GPT自動匹配
        new_mappings = auto_map_missing(
            unmapped_ids, full_path, yahoo_path, api_cfg, log_fn=self.log)

        if len(new_mappings) == 0:
            return

        # 構建新映射 dict
        new_map = {}
        for _, row in new_mappings.iterrows():
            mid = int(row['煤爐ID'])
            yid = str(row['奇摩ID'])
            yname = str(row['奇摩分類'])
            new_map[mid] = (yid, yname)

        # 回填到 DataFrame
        filled = 0
        for idx in df[unmapped_mask].index:
            reason = str(df.at[idx, '_filter_reason'])
            m = re.search(r'未映射分類\((\d+)\)', reason)
            if not m:
                continue
            sid = int(m.group(1))
            if sid in new_map:
                yid, yname = new_map[sid]
                df.at[idx, '拍賣類別'] = yid
                df.at[idx, '拍賣類別名稱'] = yname
                # 清除未映射標記
                df.at[idx, '_filter_reason'] = reason.replace(f'未映射分類({sid})', '').strip('|').strip()
                filled += 1

        if filled:
            self.log(f"自動映射回填: {filled} 行從「未映射」變為「合格」")

        # 輸出新映射文件 (追加模式，與 test 同目錄)
        output_dir = self.config.get('output_dir', '')
        if not output_dir:
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        os.makedirs(output_dir, exist_ok=True)

        out_name = self.config.get('output_name', 'test')
        map_file = os.path.join(output_dir, f'{out_name}_新增映射.xlsx')

        # 追加到已有文件
        if os.path.exists(map_file):
            try:
                existing = pd.read_excel(map_file, engine='openpyxl')
                if list(existing.columns) == list(new_mappings.columns):
                    new_mappings = pd.concat([existing, new_mappings], ignore_index=True)
                    new_mappings = new_mappings.drop_duplicates(subset=['煤爐ID'])
                    self.log(f"追加到已有映射文件 (去重後 {len(new_mappings)} 條)")
            except Exception:
                pass

        new_mappings.to_excel(map_file, index=False)
        self.log(f"✓ 新增映射已保存: {map_file}")

    # ─── AI過濾橋接 ───
    def _run_ai_filter(self, df):
        try:
            ai_dir = self.config['paths'].get('ai_filter_dir', '')
            if not ai_dir or not os.path.isdir(ai_dir):
                self.log(f"[跳過] AI過濾工具目錄不存在: {ai_dir}")
                return
            self.log("[提示] AI過濾功能尚在訓練中, 暫時跳過")
        except Exception as e:
            self.log(f"[錯誤] AI過濾失敗: {e}")

    # ─── 文件名衝突處理 ───
    @staticmethod
    def _resolve_path(filepath: str, conflict_mode: str) -> str:
        """
        處理輸出文件名衝突 (非追加模式用)
        conflict_mode:
          'rename'    → 自動加 _2, _3... 後綴
          'append'    → 由 _save_or_append 處理, 這裡直接返回原路徑
          'timestamp' → 文件名後加時間戳
        """
        if conflict_mode == 'append' or not os.path.exists(filepath):
            return filepath

        base, ext = os.path.splitext(filepath)

        if conflict_mode == 'timestamp':
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            return f"{base}_{ts}{ext}"

        # 默認 rename: 加 _2, _3...
        n = 2
        while os.path.exists(f"{base}_{n}{ext}"):
            n += 1
        return f"{base}_{n}{ext}"

    def _save_or_append(self, new_df: pd.DataFrame, filepath: str,
                        conflict_mode: str) -> str:
        """
        保存DataFrame到Excel, 追加模式下:
        - 文件不存在 → 直接寫入
        - 文件存在 + 表頭一致 → 追加行數據
        - 文件存在 + 表頭不一致 → 自動加序號另存
        返回實際保存的路徑
        """
        if conflict_mode != 'append' or not os.path.exists(filepath):
            new_df.to_excel(filepath, index=False)
            return filepath

        # 追加模式: 檢查現有文件的表頭
        try:
            existing_df = pd.read_excel(filepath, engine='openpyxl', nrows=0)
            existing_cols = list(existing_df.columns)
            new_cols = list(new_df.columns)

            if existing_cols == new_cols:
                # 表頭一致 → 讀取全部數據, 追加新數據, 重新保存
                existing_df = pd.read_excel(filepath, engine='openpyxl')
                old_count = len(existing_df)
                combined = pd.concat([existing_df, new_df], ignore_index=True)
                combined.to_excel(filepath, index=False)
                self.log(f"  [追加] 原有{old_count}行 + 新增{len(new_df)}行 = 共{len(combined)}行")
                return filepath
            else:
                # 表頭不一致 → 降級為 rename 模式
                self.log(f"  [追加] 表頭不一致, 自動另存新文件")
                base, ext = os.path.splitext(filepath)
                n = 2
                while os.path.exists(f"{base}_{n}{ext}"):
                    n += 1
                new_path = f"{base}_{n}{ext}"
                new_df.to_excel(new_path, index=False)
                return new_path

        except Exception as e:
            self.log(f"  [追加] 讀取現有文件失敗({e}), 直接覆蓋")
            new_df.to_excel(filepath, index=False)
            return filepath

    # ─── 輸出拆分與導出 ───
    def _split_and_export(self, df: pd.DataFrame, input_path: str,
                          source_type: str) -> dict:
        """拆分為合格/不合格/未分類, 導出Excel"""
        output_dir = self.config.get('output_dir', '')
        if not output_dir:
            output_dir = os.path.join(os.path.dirname(input_path), 'output')
        os.makedirs(output_dir, exist_ok=True)

        # 自定義文件名 (默認 test)
        out_name = self.config.get('output_name', '').strip()
        if not out_name:
            out_name = 'test'

        conflict_mode = self.config.get('output_conflict', 'rename')

        # 分離
        has_reason = df['_filter_reason'].fillna('').str.strip().astype(bool)
        is_uncat = df['_filter_reason'].str.contains('未分類|未映射', na=False)

        good_df = df[~has_reason].copy()
        bad_df = df[has_reason & ~is_uncat].copy()
        uncat_df = df[is_uncat].copy()

        # 輸出列整理
        out_cols = [c for c in OUTPUT_COLUMNS if c in df.columns]

        # 導出合格 → {out_name}.xlsx
        good_raw = os.path.join(output_dir, f'{out_name}.xlsx')
        good_path = self._resolve_path(good_raw, conflict_mode)
        if len(good_df) > 0:
            good_out = good_df[out_cols].copy().fillna('')
            good_path = self._save_or_append(good_out, good_path, conflict_mode)
            self._format_output_xlsx(good_path)
            self.log(f"✓ 合格商品: {good_path} ({len(good_df)} 行)")

        # 導出不合格 → {out_name}_不合格.xlsx
        bad_raw = os.path.join(output_dir, f'{out_name}_不合格.xlsx')
        bad_path = self._resolve_path(bad_raw, conflict_mode)
        if len(bad_df) > 0:
            bad_out = bad_df[out_cols + ['_filter_reason']].copy().fillna('')
            bad_out = bad_out.rename(columns={'_filter_reason': '過濾原因'})
            bad_path = self._save_or_append(bad_out, bad_path, conflict_mode)
            self._format_output_xlsx(bad_path)
            self.log(f"✗ 不合格商品: {bad_path} ({len(bad_df)} 行)")

        # 導出未分類 → {out_name}_未分類.xlsx
        uncat_raw = os.path.join(output_dir, f'{out_name}_未分類.xlsx')
        uncat_path = self._resolve_path(uncat_raw, conflict_mode)
        if len(uncat_df) > 0:
            uncat_out = uncat_df[['標題', '商品條碼', '_source_category_id', '_filter_reason']].copy().fillna('')
            uncat_out = uncat_out.rename(columns={
                '_source_category_id': '來源分類ID',
                '_filter_reason': '過濾原因'
            })
            uncat_path = self._save_or_append(uncat_out, uncat_path, conflict_mode)
            self._format_output_xlsx(uncat_path)
            self.log(f"? 未分類商品: {uncat_path} ({len(uncat_df)} 行)")

        return {
            'good_path': good_path if len(good_df) > 0 else '',
            'bad_path': bad_path if len(bad_df) > 0 else '',
            'uncat_path': uncat_path if len(uncat_df) > 0 else '',
            'good_count': len(good_df),
            'bad_count': len(bad_df),
            'uncat_count': len(uncat_df),
        }

    def _format_output_xlsx(self, xlsx_path: str):
        """格式化輸出Excel: 清除nan, 數字列轉純數字, 自動列寬"""
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            headers = [c.value for c in ws[1]]

            # 先清除所有殘留的 "nan" 字串
            for r in range(2, ws.max_row + 1):
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None and str(cell.value).strip().lower() == 'nan':
                        cell.value = None

            # 所有可能含數字ID的列 → 轉為純整數
            num_cols = ['拍賣類別', '商品條碼', '來源分類ID']
            for col_name in num_cols:
                if col_name in headers:
                    cidx = headers.index(col_name) + 1
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=cidx)
                        val = cell.value
                        if val is not None:
                            s = str(val).strip()
                            if s and s not in ('', 'nan', 'None'):
                                try:
                                    cell.value = int(float(s))
                                    cell.number_format = '0'
                                except (ValueError, OverflowError):
                                    pass

            # 自動列寬 (根據表頭和內容)
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1] or ''))
                for r in range(2, min(ws.max_row + 1, 52)):  # 取樣前50行
                    val = ws.cell(row=r, column=col_idx).value
                    if val is not None:
                        vlen = len(str(val))
                        if vlen > max_len:
                            max_len = vlen
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

            wb.save(xlsx_path)
            wb.close()
        except Exception as e:
            self.log(f"[警告] 格式化輸出文件失敗: {e}")


def load_config(config_path: str = None) -> dict:
    """載入配置文件, 不存在或損壞時返回默認配置"""
    if config_path is None:
        config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {
            'source_type': 'mercari',
            'last_input_file': '',
            'output_dir': '',
            'output_name': 'test',
            'output_conflict': 'rename',
            'process_mode': 'raw',
            'steps': {
                'translate': {'enabled': False, 'auto_for_mercari': True},
                'seo': {'enabled': False},
                'category': {'enabled': True},
                'replace': {'enabled': True},
                'keyword': {'enabled': True},
                'price': {'enabled': True},
                'defaults': {'enabled': True},
                'ai_filter': {'enabled': False},
                'dedup': {'enabled': True},
                'desc_append': {'enabled': False},
            },
            'paths': {},
            'price': {
                'mercari_divisor': 20,
                'min_price': 15,
                'max_price': 80000,
                'floor_price': 700,
                'tiers': [
                    {'min': 10000, 'max': 999999, 'multiplier': 8},
                    {'min': 5000, 'max': 9999, 'multiplier': 9},
                    {'min': 2000, 'max': 4999, 'multiplier': 9.5},
                    {'min': 500, 'max': 1999, 'multiplier': 10},
                    {'min': 100, 'max': 499, 'multiplier': 12},
                    {'min': 70, 'max': 99, 'multiplier': 13},
                    {'min': 50, 'max': 69, 'multiplier': 14},
                    {'min': 40, 'max': 49, 'multiplier': 17},
                ],
            },
            'defaults': {
                'mercari': {},
                'goofish': {},
            },
            'api': {},
        }


def save_config(config: dict, config_path: str = None):
    """保存配置文件"""
    if config_path is None:
        config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
